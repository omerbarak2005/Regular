#!/usr/bin/env python3
"""
Formats qsr_outreach.csv into a clean, color-coded Excel workbook.
Usage: python format_excel.py [--input qsr_outreach.csv] [--output QSR_Outreach.xlsx]
"""

import csv
import argparse
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Palette ────────────────────────────────────────────────────────────────────
WHITE       = "FFFFFF"
BLACK       = "000000"

# Header
HDR_BG      = "1F3864"   # deep navy
HDR_FG      = "FFFFFF"

# Priority rows
P1_BG       = "E2EFDA"   # soft green
P2_BG       = "FFF2CC"   # soft amber
P3_BG       = "F2F2F2"   # light gray
P1_ACC      = "375623"   # dark green text accent
P2_ACC      = "7F6000"   # dark amber text accent

# Firm type badges (Priority col)
TYPE_PE         = "D6E4F7"   # steel blue
TYPE_STRATEGIC  = "FCE4D6"   # salmon
TYPE_FO         = "E2D4F0"   # lavender
TYPE_WEALTH     = "D4EDDA"   # mint

# Email confidence
CONF_CONFIRMED  = "C6EFCE"   # green
CONF_PROBABLE   = "FFEB9C"   # yellow
CONF_FORMAT     = "FFC7CE"   # red-pink

# Alternating rows (subtle)
ALT_LIGHT   = "FFFFFF"
ALT_DARK    = "F7F9FC"

# Border color
BORDER_CLR  = "BDD7EE"

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=BLACK, size=10, italic=False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def _border() -> Border:
    side = Side(style="thin", color=BORDER_CLR)
    return Border(left=side, right=side, top=side, bottom=side)

def _align(wrap=False, h="left", v="center") -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Column config ──────────────────────────────────────────────────────────────
# (header_label, csv_key, width, wrap, h_align)
COLUMNS = [
    ("Priority",          "Priority",          8,   False, "center"),
    ("Date",              "Date",              10,   False, "center"),
    ("Company",           "Company",           28,   False, "left"),
    ("Contact",           "Contact",           24,   False, "left"),
    ("Title",             "Title",             30,   True,  "left"),
    ("Email",             "Email",             36,   False, "left"),
    ("Phone",             "Phone",             16,   False, "center"),
    ("Firm Type",         "Firm Type",         14,   False, "center"),
    ("Email Confidence",  "Email Confidence",  18,   False, "center"),
    ("Activity Type",     "Activity Type",     18,   False, "center"),
    ("NDA Status",        "NDA Status",        12,   False, "center"),
    ("Status",            "Status After",      26,   False, "left"),
    ("Logged By",         "Logged By",         12,   False, "center"),
    ("Next Steps",        "Next Steps",        24,   True,  "left"),
    ("Email Subject",     "Email Subject",     40,   True,  "left"),
    ("Email Body",        "Email Body",        60,   True,  "left"),
    ("Notes",             "Notes",             40,   True,  "left"),
]


def priority_bg(row: dict) -> str:
    p = str(row.get("Priority", "2"))
    return {
        "1": P1_BG,
        "2": P2_BG,
        "3": P3_BG,
    }.get(p, ALT_LIGHT)

def priority_font(row: dict) -> Font:
    p = str(row.get("Priority", "2"))
    if p == "1":
        return _font(bold=True, color=P1_ACC)
    elif p == "2":
        return _font(bold=False, color=P2_ACC)
    return _font(color="444444")

def confidence_fill(val: str) -> PatternFill:
    v = val.lower()
    if "confirmed" in v:
        return _fill(CONF_CONFIRMED)
    elif "probable" in v:
        return _fill(CONF_PROBABLE)
    return _fill(CONF_FORMAT)

def type_fill(val: str) -> PatternFill:
    v = val.lower()
    if "strategic" in v:
        return _fill(TYPE_STRATEGIC)
    elif "family" in v:
        return _fill(TYPE_FO)
    elif "wealth" in v:
        return _fill(TYPE_WEALTH)
    return _fill(TYPE_PE)

def priority_badge_fill(val: str) -> PatternFill:
    v = str(val)
    return {
        "1": _fill("375623"),
        "2": _fill("806000"),
        "3": _fill("595959"),
    }.get(v, _fill("595959"))

def priority_badge_font(val: str) -> Font:
    return Font(bold=True, color=WHITE, size=10, name="Calibri")


def build_excel(rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "QSR Outreach Pipeline"

    # ── Header row ──────────────────────────────────────────────────────────────
    for col_idx, (label, _, width, wrap, halign) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = _fill(HDR_BG)
        cell.font      = _font(bold=True, color=HDR_FG, size=11)
        cell.alignment = _align(wrap=False, h="center", v="center")
        cell.border    = _border()
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 28

    # ── Data rows ────────────────────────────────────────────────────────────────
    for row_idx, row in enumerate(rows, start=2):
        bg      = priority_bg(row)
        row_h   = priority_font(row)
        is_dark = (row_idx % 2 == 0)

        for col_idx, (_, csv_key, _, wrap, halign) in enumerate(COLUMNS, start=1):
            val  = row.get(csv_key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border    = _border()
            cell.alignment = _align(wrap=wrap, h=halign, v="top")

            # Base row color
            cell.fill = _fill(bg if not is_dark else _darken(bg))
            cell.font = row_h

            # Override specific columns with semantic coloring
            label = COLUMNS[col_idx - 1][0]

            if label == "Priority":
                cell.fill      = priority_badge_fill(val)
                cell.font      = priority_badge_font(val)
                cell.alignment = _align(h="center", v="center")

            elif label == "Email Confidence":
                cell.fill = confidence_fill(str(val))
                cell.font = _font(bold=("confirmed" in str(val).lower()), size=9)
                cell.alignment = _align(h="center", v="center")

            elif label == "Firm Type":
                cell.fill      = type_fill(str(val))
                cell.font      = _font(size=9)
                cell.alignment = _align(h="center", v="center")

            elif label == "Email":
                cell.font = Font(
                    color="1155CC", underline="single",
                    size=10, name="Calibri",
                    bold=("confirmed" in str(row.get("Email Confidence", "")).lower())
                )

            elif label == "Status":
                cell.font = _font(italic=True, color="444444", size=9)

            elif label in ("Email Subject", "Email Body"):
                cell.font = _font(color="1F3864", size=9)

            elif label == "Notes":
                cell.font = _font(italic=True, color="595959", size=9)

        # Row height — taller for rows with body text
        ws.row_dimensions[row_idx].height = 56 if str(row.get("Priority","")) == "1" else 46

    # ── Freeze header + filter ──────────────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Legend sheet ────────────────────────────────────────────────────────────
    leg = wb.create_sheet("Legend")
    leg.column_dimensions["A"].width = 22
    leg.column_dimensions["B"].width = 36

    legend_items = [
        ("PRIORITY",          "",             HDR_BG,         HDR_FG,   True),
        ("P1 — HOT",          "Send immediately",          "375623", WHITE, True),
        ("P2 — WARM",         "Send second wave",          "806000", WHITE, False),
        ("P3 — COLD",         "Family office / longer shot","595959", WHITE, False),
        ("",                  "",             WHITE,          BLACK,    False),
        ("EMAIL CONFIDENCE",  "",             HDR_BG,         HDR_FG,   True),
        ("✓ Confirmed",       "Verified via multiple sources",  CONF_CONFIRMED, P1_ACC, True),
        ("~ Probable",        "Based on firm email format",     CONF_PROBABLE,  P2_ACC, False),
        ("? Format only",     "Unverified — check before send", CONF_FORMAT,    "9C0006",False),
        ("",                  "",             WHITE,          BLACK,    False),
        ("FIRM TYPE",         "",             HDR_BG,         HDR_FG,   True),
        ("PE",                "Private equity firm",            TYPE_PE,        "1F3864",False),
        ("Strategic",         "Operator / public co acquirer",  TYPE_STRATEGIC, "843C0C",False),
        ("Family Office",     "Family office / direct investor",TYPE_FO,        "5B2D8E",False),
        ("Wealth",            "Wealth manager / RIA",           TYPE_WEALTH,    "1A5C38",False),
    ]

    leg.cell(row=1, column=1, value="QSR Outreach — Color Legend").font = _font(bold=True, color=HDR_FG, size=13)
    leg.cell(row=1, column=1).fill = _fill(HDR_BG)
    leg.merge_cells("A1:B1")
    leg.row_dimensions[1].height = 26

    for i, (label, desc, bg, fg, bold) in enumerate(legend_items, start=2):
        ca = leg.cell(row=i, column=1, value=label)
        cb = leg.cell(row=i, column=2, value=desc)
        for c in (ca, cb):
            c.fill      = _fill(bg)
            c.font      = _font(bold=bold, color=fg, size=10)
            c.border    = _border()
            c.alignment = _align(h="left", v="center")
        leg.row_dimensions[i].height = 20

    # ── Summary sheet ────────────────────────────────────────────────────────────
    smry = wb.create_sheet("Summary")
    smry.column_dimensions["A"].width = 32
    smry.column_dimensions["B"].width = 14

    p1 = [r for r in rows if str(r.get("Priority","")) == "1"]
    p2 = [r for r in rows if str(r.get("Priority","")) == "2"]
    p3 = [r for r in rows if str(r.get("Priority","")) == "3"]
    confirmed = [r for r in rows if "confirmed" in str(r.get("Email Confidence","")).lower()]
    pe_rows   = [r for r in rows if r.get("Firm Type","") == "pe"]
    strat     = [r for r in rows if r.get("Firm Type","") == "strategic"]
    fo        = [r for r in rows if r.get("Firm Type","") in ("family_office","wealth")]

    summary_data = [
        ("PROJECT ALPINE — DEAL SUMMARY",   "",                 HDR_BG,   HDR_FG, True,  13),
        ("Concept",       "34-Unit Drive-Thru Coffee, PNW",     P1_BG,    P1_ACC, False, 11),
        ("Revenue",       "$70M+",                              P1_BG,    P1_ACC, True,  11),
        ("Adj. EBITDA",   "$13M+",                              P1_BG,    P1_ACC, True,  11),
        ("EBITDA Margin", "~18.5%",                             P1_BG,    P1_ACC, True,  11),
        ("EV Range",      "$78M–$130M (6–10x EBITDA)",          P1_BG,    P1_ACC, False, 11),
        ("Owners",        "3 co-owners, aligned on timeline",   P1_BG,    P1_ACC, False, 11),
        ("",              "",                                   WHITE,    BLACK,  False, 11),
        ("PIPELINE SUMMARY",  "",                               HDR_BG,   HDR_FG, True,  12),
        ("Total Targets",     str(len(rows)),                   P2_BG,    P2_ACC, True,  11),
        ("Priority 1 (HOT)",  str(len(p1)),                    P1_BG,    P1_ACC, True,  11),
        ("Priority 2 (WARM)", str(len(p2)),                    P2_BG,    P2_ACC, False, 11),
        ("Priority 3 (COLD)", str(len(p3)),                    P3_BG,    "444444",False, 11),
        ("Confirmed Emails",  str(len(confirmed)),              CONF_CONFIRMED, P1_ACC, True, 11),
        ("PE Firms",          str(len(pe_rows)),                TYPE_PE,  "1F3864",False, 11),
        ("Strategic Buyers",  str(len(strat)),                  TYPE_STRATEGIC,"843C0C",False,11),
        ("Family Office/Wealth", str(len(fo)),                  TYPE_FO,  "5B2D8E",False, 11),
        ("",              "",                                   WHITE,    BLACK,  False, 11),
        ("ACTION PLAN",   "",                                   HDR_BG,   HDR_FG, True,  12),
        ("Step 1",  "Update ADVISOR in outreach.py with real contact info", P1_BG, P1_ACC, True, 10),
        ("Step 2",  "Fix 11 wrong contacts in tracker (see Notes col)",     P2_BG, P2_ACC, False,10),
        ("Step 3",  "Send Priority 1 wave (11 firms) this week",            P1_BG, P1_ACC, True, 10),
        ("Step 4",  "Follow up in 5 business days, send P2 wave",           P2_BG, P2_ACC, False,10),
        ("Step 5",  "Run --hunter-key <KEY> to verify emails",              P3_BG, "444444",False,10),
    ]

    for i, (label, val, bg, fg, bold, sz) in enumerate(summary_data, start=1):
        ca = smry.cell(row=i, column=1, value=label)
        cb = smry.cell(row=i, column=2, value=val)
        for c in (ca, cb):
            c.fill      = _fill(bg)
            c.font      = _font(bold=bold, color=fg, size=sz)
            c.border    = _border()
            c.alignment = _align(h="left", v="center", wrap=True)
        smry.row_dimensions[i].height = 22 if sz >= 12 else 18

    wb.save(output_path)
    print(f"✓ Saved formatted workbook → {output_path}")
    print(f"  Sheets: 'QSR Outreach Pipeline' | 'Summary' | 'Legend'")
    print(f"  Rows:   {len(rows)} contacts  |  {len(p1)} P1  |  {len(p2)} P2  |  {len(p3)} P3")


def _darken(hex_color: str, factor: float = 0.97) -> str:
    """Slightly darken a hex color for alternating rows."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r * factor)
    g = int(g * factor)
    b = int(b * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def main():
    parser = argparse.ArgumentParser(description="Format QSR outreach CSV into Excel")
    parser.add_argument("--input",  default="qsr_outreach.csv",  help="Input CSV file")
    parser.add_argument("--output", default="QSR_Outreach_Pipeline.xlsx", help="Output Excel file")
    args = parser.parse_args()

    with open(args.input, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    build_excel(rows, args.output)


if __name__ == "__main__":
    main()
