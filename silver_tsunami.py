#!/usr/bin/env python3
"""
SILVER TSUNAMI — Boomer Business Sourcing Engine
Boutique IB Deal Flow Machine | Lower Middle Market M&A

Three modes:
  1. PPP data scan   → python silver_tsunami.py --ppp ppp_data.csv --naics fire_safety
  2. Manual prospect → python silver_tsunami.py --manual
  3. Full export     → python silver_tsunami.py --ppp ppp_data.csv --output deals.xlsx

PPP data (free, public):
  Download from https://data.sba.gov/dataset/ppp-foia
  Pick your state CSV files, merge them, feed to --ppp flag.

How to estimate revenue from PPP loan:
  PPP loan = 2.5x monthly payroll
  Service biz payroll ≈ 35-45% of revenue
  So: Revenue ≈ (loan / 2.5) / 0.4 * 12
  Sweet spot loan size: $300K–$5M = ~$9M–$60M revenue
"""

import csv, sys, json, math, argparse, os
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Dict
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = date.today().strftime("%-m/%-d/%y")

# ═══════════════════════════════════════════════════════════════════════════════
#  YOUR FIRM  —  fill these in
# ═══════════════════════════════════════════════════════════════════════════════
FIRM = {
    "name":    "[Your Firm Name]",
    "banker":  "[Your Name]",
    "title":   "Managing Director",
    "email":   "you@yourfirm.com",
    "phone":   "(212) XXX-XXXX",
    "city":    "New York, NY",
    "tagline": "Boutique M&A Advisory | Lower Middle Market",
}

# ═══════════════════════════════════════════════════════════════════════════════
#  NICHE INDUSTRY DATABASE
#  These are the overlooked, under-brokered sectors where boomers are retiring
#  and PE firms are paying real multiples with zero competition from big banks.
# ═══════════════════════════════════════════════════════════════════════════════
INDUSTRIES: Dict[str, dict] = {

    "fire_safety": {
        "name": "Fire & Life Safety Inspection",
        "description": (
            "Code-mandated annual inspections, recurring contracts, zero customer churn. "
            "Pye-Barker did 57 acquisitions in 2024 alone. Most owners are 60+ running "
            "a shop they built from scratch, never thought about selling."
        ),
        "why_pe_loves_it": [
            "100% recurring revenue from code-mandated inspections",
            "Customer retention 95%+ — switching costs are regulatory nightmares",
            "Fragmented: 15,000+ independent operators in the US",
            "Essential service — recessions don't stop fire code",
            "Low capex: vans + inspection equipment, no factories",
        ],
        "naics_codes": ["561621", "238290", "238220"],
        "ebitda_multiple": "7–10x",
        "deal_size": "$5M–$60M EV",
        "boomer_density": "EXTREME",
        "competition_for_deal": "LOW — brokers don't know this space",
        "buyers": [
            {"name": "Pye-Barker Fire & Safety",     "type": "strategic_rollup", "contact": "M&A Team",          "email": "acquisitions@pyebarkerfs.com",   "note": "Did 57 deals in 2024 — most active buyer on planet"},
            {"name": "APi Group",                    "type": "public_strategic", "contact": "Corp Dev",           "email": "investor.relations@apigroup.com", "note": "NYSE: APG, owns Elevated Facility Services"},
            {"name": "Convergint Technologies",      "type": "pe_rollup",        "contact": "M&A",                "email": "info@convergint.com",            "note": "PE-backed, very active in life safety"},
            {"name": "Johnson Controls",             "type": "strategic",        "contact": "Corp Dev",           "email": "corpdev@jci.com",                "note": "Strategic acquirer, pays premium for book of business"},
            {"name": "Koorsen Fire & Security",      "type": "pe_rollup",        "contact": "Acquisitions",       "email": "info@koorsen.com",               "note": "Midwest-based rollup, expanding"},
            {"name": "Alarm Capital Alliance",       "type": "pe_rollup",        "contact": "M&A",                "email": "info@alarmcapital.com",          "note": "Fire + security rollup"},
        ],
    },

    "elevator": {
        "name": "Elevator & Escalator Inspection/Maintenance",
        "description": (
            "State-licensed, code-required inspections and maintenance. Otis/Schindler/KONE "
            "dominate new installations but independent shops handle 40% of maintenance. "
            "APi Group paid $570M for Elevated Facility Services in 2024 — local operators "
            "worth $5M–$30M get zero attention from bankers."
        ),
        "why_pe_loves_it": [
            "State licensing creates a near-impenetrable moat",
            "Multi-year maintenance contracts, near-zero churn",
            "Elevator inspections mandated by law — recession proof",
            "Aging building stock = growing installed base",
            "Consolidators paying 8-11x EBITDA for geographic coverage",
        ],
        "naics_codes": ["238290", "811310", "238220"],
        "ebitda_multiple": "8–11x",
        "deal_size": "$5M–$40M EV",
        "boomer_density": "EXTREME",
        "competition_for_deal": "VERY LOW — almost no bankers touch this",
        "buyers": [
            {"name": "APi Group / Elevated Facility", "type": "public_strategic", "contact": "Corp Dev",    "email": "investor.relations@apigroup.com",   "note": "Paid $570M for Elevated — now buying add-ons"},
            {"name": "3Phase Elevator",               "type": "pe_rollup",        "contact": "Acquisitions","email": "info@3phaseelevator.com",           "note": "Active independent rollup"},
            {"name": "Thyssenkrupp Elevator (TK)",    "type": "strategic",        "contact": "Corp Dev",    "email": "info@thyssenkrupp-elevator.com",    "note": "PE-backed (Advent/Cinven), buying service books"},
            {"name": "Schindler Group",               "type": "strategic",        "contact": "M&A USA",     "email": "us.info@schindler.com",             "note": "Strategic — pays for market share"},
            {"name": "Elevator Service Company",      "type": "pe_rollup",        "contact": "Acquisitions","email": "acquisitions@elevatorservice.com",  "note": "Active independent rollup"},
        ],
    },

    "calibration": {
        "name": "Calibration & Metrology Labs",
        "description": (
            "Every manufacturer, pharma company, aerospace firm, and hospital must "
            "calibrate their measurement equipment to ISO/FDA standards. Transcat "
            "(NASDAQ: TRNS) just paid $79M for a 7-lab network doing $25M revenue. "
            "They've done 26 acquisitions total. Local lab with 3-5 engineers and "
            "20-year customer relationships = perfect target. Zero bankers calling."
        ),
        "why_pe_loves_it": [
            "Regulated industries (pharma, aerospace, medical) = non-negotiable recurring spend",
            "ISO 17025 accreditation takes 2 years — massive barrier to entry",
            "Lab customers stay for decades, switching requires requalification",
            "High margin: technical expertise, minimal physical assets",
            "ESG + supply chain resilience driving more domestic testing",
        ],
        "naics_codes": ["811219", "541380", "541330"],
        "ebitda_multiple": "7–12x",
        "deal_size": "$5M–$50M EV",
        "boomer_density": "HIGH",
        "competition_for_deal": "VERY LOW — extremely niche, zero generalist bankers",
        "buyers": [
            {"name": "Transcat Inc.",          "type": "public_rollup",   "contact": "Lee Rudow, CEO",     "email": "lrudow@transcat.com",            "note": "NASDAQ: TRNS, 26 acquisitions, most active buyer"},
            {"name": "Trescal",                "type": "pe_rollup",       "contact": "Tom Mathews, CEO USA","email": "info@trescal.com",               "note": "European PE-backed, aggressively expanding US"},
            {"name": "Integrated Defense Tech","type": "pe_rollup",       "contact": "M&A",                "email": "info@idtus.com",                 "note": "Defense-focused calibration rollup"},
            {"name": "Intertek",               "type": "public_strategic","contact": "Corp Dev",           "email": "corpdev@intertek.com",           "note": "London-listed, buying US cal labs"},
            {"name": "Bureau Veritas",         "type": "public_strategic","contact": "Corp Dev",           "email": "corpdev@bureauveritas.com",      "note": "French-listed, active US acquirer"},
            {"name": "Team Precision",         "type": "pe_rollup",       "contact": "Acquisitions",       "email": "info@teamprecision.com",         "note": "Smaller rollup, moves fast"},
        ],
    },

    "specialty_recycling": {
        "name": "Specialty / Regulated Recycling",
        "description": (
            "E-waste, battery, fluorescent lamp, pharmaceutical, and industrial "
            "chemical recycling. EPA compliance is non-negotiable — every hospital, "
            "manufacturer, and school district needs a licensed vendor. ESG pressure "
            "from corporate clients making this a must-have service. "
            "Most operators are founder-owned 60+ year-olds who built their business "
            "on relationships, with zero succession plan."
        ),
        "why_pe_loves_it": [
            "Regulatory tailwind: EPA mandates + corporate ESG commitments",
            "Licensing creates near-impenetrable local monopolies",
            "Battery recycling exploding with EV adoption",
            "Long-term contracts with hospitals, manufacturers, municipalities",
            "Clean Earth, US Ecology, Retriev all buying aggressively",
        ],
        "naics_codes": ["562920", "562219", "562211", "562998"],
        "ebitda_multiple": "7–11x",
        "deal_size": "$5M–$50M EV",
        "boomer_density": "HIGH",
        "competition_for_deal": "LOW",
        "buyers": [
            {"name": "Clean Earth (Harsco)",     "type": "pe_rollup",       "contact": "Corp Dev",          "email": "info@cleanearthcapital.com",     "note": "Most active specialty waste buyer"},
            {"name": "Retriev Technologies",     "type": "pe_rollup",       "contact": "Acquisitions",      "email": "info@retriev.com",               "note": "Battery/lithium-ion specialist, Tesla partner"},
            {"name": "US Ecology (Republic)",    "type": "strategic",       "contact": "Corp Dev",          "email": "corpdev@republic.com",           "note": "Republic Services subsidiary, buying specialty waste"},
            {"name": "Covanta Energy",           "type": "strategic",       "contact": "Corp Dev",          "email": "info@covanta.com",               "note": "Waste-to-energy, strategic acquirer"},
            {"name": "Veolia Environmental",     "type": "public_strategic","contact": "M&A USA",           "email": "na@veolia.com",                  "note": "French-listed, very active US environmental services"},
            {"name": "Enviro Star",              "type": "pe_rollup",       "contact": "Acquisitions",      "email": "info@envirostarinc.com",         "note": "Environmental services rollup"},
        ],
    },

    "portable_sanitation": {
        "name": "Portable Sanitation / Liquid Waste",
        "description": (
            "Portable toilet rental, septic pumping, grease trap service, hydrovac. "
            "Route-based, recession-proof, zero tech disruption risk. "
            "United Site Services (backed by GI Partners) is the national rollup "
            "but thousands of regional operators exist — most run by a 65-year-old "
            "who bought his first truck in 1985."
        ),
        "why_pe_loves_it": [
            "Route density = massive margin improvement post-acquisition",
            "Long-term contracts with construction, events, municipalities",
            "Regulatory compliance drives recurring grease trap work",
            "Hydrovac demand growing with infrastructure bill spending",
            "Classic rollup math: buy 10 locals, fire duplicate overhead, 3x EBITDA",
        ],
        "naics_codes": ["562991", "562119", "238310", "532490"],
        "ebitda_multiple": "6–9x",
        "deal_size": "$5M–$40M EV",
        "boomer_density": "EXTREME",
        "competition_for_deal": "LOW — nobody wants to talk about toilets",
        "buyers": [
            {"name": "United Site Services",    "type": "pe_rollup",   "contact": "Corp Dev / M&A",  "email": "info@unitedsiteservices.com",   "note": "GI Partners-backed, most aggressive buyer"},
            {"name": "Clean Harbors",           "type": "public",      "contact": "Corp Dev",        "email": "corpdev@cleanharbors.com",      "note": "NYSE: CLH, buying liquid waste books"},
            {"name": "Synagro Technologies",    "type": "pe_rollup",   "contact": "Acquisitions",    "email": "info@synagro.com",              "note": "Biosolids + liquid waste specialist"},
            {"name": "ZAGO (Canada)",           "type": "pe_rollup",   "contact": "M&A",             "email": "info@zago.ca",                  "note": "Cross-border rollup, active in US"},
            {"name": "New Market Waste",        "type": "pe_rollup",   "contact": "Acquisitions",    "email": "info@newmarketwaste.com",       "note": "Grease trap specialist rollup"},
        ],
    },

    "document_destruction": {
        "name": "Secure Document Destruction / Shredding",
        "description": (
            "HIPAA, FACTA, and state privacy laws require certified document destruction. "
            "Every law firm, hospital, bank, and accounting firm is a recurring customer. "
            "Shred-it and Iron Mountain dominate nationally but thousands of regional "
            "operators serve local markets — owner has 500 clients paying monthly, "
            "thinks his company is worth 2x revenue. It's worth 6-8x EBITDA."
        ),
        "why_pe_loves_it": [
            "HIPAA/FACTA compliance drives non-negotiable recurring contracts",
            "COC (Certificate of Destruction) builds switching costs",
            "Route density = massive margin expansion post-rollup",
            "Digital transformation actually increasing demand (fear of data breaches)",
            "Shred-it doing $2B+ — but they only buy 50+ truck fleets",
        ],
        "naics_codes": ["561439", "562111", "561499"],
        "ebitda_multiple": "6–9x",
        "deal_size": "$3M–$30M EV",
        "boomer_density": "HIGH",
        "competition_for_deal": "MEDIUM — some brokers active here",
        "buyers": [
            {"name": "Shred-it (Stericycle/WM)",  "type": "strategic",   "contact": "Corp Dev",    "email": "corpdev@shred-it.com",         "note": "Largest buyer, wants 50+ truck fleets only"},
            {"name": "Iron Mountain",              "type": "public",      "contact": "Corp Dev",    "email": "corpdev@ironmountain.com",     "note": "NYSE: IRM, buys smaller regional operators"},
            {"name": "Proshred Security",          "type": "pe_rollup",   "contact": "Franchise/M&A","email": "info@proshred.com",           "note": "Franchise model buying independent operators"},
            {"name": "Greenway Shredding",         "type": "pe_rollup",   "contact": "Acquisitions","email": "info@greenwayshredding.com",   "note": "Regional rollup, moves fast"},
            {"name": "Secure Shredding",           "type": "pe_rollup",   "contact": "M&A",         "email": "info@secureshredding.com",    "note": "Smaller regional buyer, pays fair multiples"},
        ],
    },

    "funeral_homes": {
        "name": "Funeral Homes & Death Care",
        "description": (
            "The most recession-proof business on earth. Death is not optional. "
            "Average funeral home has been family-owned for 3 generations — owner is 68, "
            "kids don't want the business, zero succession plan. "
            "Park Lawn (TSX: PLC) and SCI are buying everything. "
            "Multiples look low (4-6x) but the cash flows are Swiss-clock predictable."
        ),
        "why_pe_loves_it": [
            "Literally recession-proof — death rate is independent of GDP",
            "Pre-need contracts (prepaid funerals) = massive recurring revenue",
            "Community trust moat — families use same funeral home for generations",
            "Real estate often included (funeral home owns the building)",
            "Aging US population = secular tailwind for 30+ years",
        ],
        "naics_codes": ["812210", "812220"],
        "ebitda_multiple": "5–8x",
        "deal_size": "$3M–$30M EV",
        "boomer_density": "EXTREME — owner often is literally a boomer",
        "competition_for_deal": "MEDIUM",
        "buyers": [
            {"name": "Service Corporation International","type": "public",      "contact": "Corp Dev",       "email": "corpdev@sci-us.com",           "note": "NYSE: SCI, largest death care company"},
            {"name": "Park Lawn Corporation",           "type": "public_rollup","contact": "J. Bradley Green","email": "info@parklawn.com",           "note": "TSX: PLC, most aggressive buyer right now"},
            {"name": "Carriage Services",               "type": "public",       "contact": "Corp Dev",       "email": "corpdev@carriageservices.com", "note": "NYSE: CSV, active mid-size acquirer"},
            {"name": "Foundation Partners Group",       "type": "pe_rollup",    "contact": "Acquisitions",   "email": "info@foundationpartners.com", "note": "Backed by Stonehenge Partners"},
            {"name": "NorthStar Memorial Group",        "type": "pe_rollup",    "contact": "Corp Dev",       "email": "info@northstarmemorial.com",  "note": "PE-backed, regional rollup"},
            {"name": "Loewen Group (InvoCare)",         "type": "pe_rollup",    "contact": "M&A",            "email": "info@invocare.com.au",        "note": "Australian-listed, buying US operators"},
        ],
    },

    "insurance_agencies": {
        "name": "Independent Insurance Agencies",
        "description": (
            "The hottest rollup market right now. PE firms are paying 8-12x commission "
            "revenue for books of business — and the owner thinks it's worth 1-2x. "
            "Every independent agent who's 60+ and built a $2M-$10M P&C book over "
            "30 years is a target. 36,000+ independent agencies in the US. "
            "BroadStreet Partners is calling every one of them."
        ),
        "why_pe_loves_it": [
            "Commission revenue is pure recurring — clients renew annually forever",
            "No accounts receivable — commissions paid by carriers",
            "Owner-operator = massive underinvestment in technology/staff = value creation",
            "Cross-sell opportunity is enormous once you have the client",
            "37,000 independent agencies = unlimited M&A pipeline",
        ],
        "naics_codes": ["524210", "524110", "524130"],
        "ebitda_multiple": "8–12x",
        "deal_size": "$5M–$50M EV",
        "boomer_density": "EXTREME",
        "competition_for_deal": "HIGH — BroadStreet, Acrisure both very active",
        "buyers": [
            {"name": "BroadStreet Partners",   "type": "pe_rollup",   "contact": "Ryan Dobratz",     "email": "rdobratz@broadstreetpartners.com", "note": "Most active LMM buyer, hyper-aggressive"},
            {"name": "Acrisure",               "type": "pe_rollup",   "contact": "Greg Williams CEO","email": "info@acrisure.com",               "note": "Largest PE-backed agency, $23B AUM clients"},
            {"name": "AssuredPartners",        "type": "pe_rollup",   "contact": "Corp Dev",         "email": "info@assuredpartners.com",        "note": "GTCR-backed, very active"},
            {"name": "Hub International",      "type": "pe_rollup",   "contact": "Corp Dev",         "email": "info@hubinternational.com",       "note": "Hellman & Friedman, paying 10-12x"},
            {"name": "Inszone Insurance",      "type": "pe_rollup",   "contact": "Chris Walters CEO","email": "cwalters@inszone.com",            "note": "Fast-growing Western US rollup"},
            {"name": "Patriot Growth Insurance","type": "pe_rollup",  "contact": "Corp Dev",         "email": "info@patriotgrowth.com",          "note": "Summit Partners-backed, aggressive"},
        ],
    },

    "cold_storage": {
        "name": "Cold Storage / Refrigerated Warehousing",
        "description": (
            "Food supply chain is non-negotiable. Lineage Logistics (KKR-backed) "
            "went public at $18B in 2024 — the largest REIT IPO in history. "
            "Regional cold storage operators doing $5M-$30M revenue are completely "
            "under the radar. Owner built one 40,000 sq ft freezer warehouse in 1992, "
            "now has 5 customers who can't live without him."
        ),
        "why_pe_loves_it": [
            "Infrastructure-like cash flows — food companies have nowhere else to go",
            "New cold storage construction costs $150-$250/sq ft — huge replacement moat",
            "E-commerce grocery + meal kit explosion driving demand surge",
            "Onshoring pharma cold chain (vaccines, biologics) = new demand",
            "Lineage/Americold pay premium for geographic coverage",
        ],
        "naics_codes": ["493120", "493110", "488991"],
        "ebitda_multiple": "8–13x",
        "deal_size": "$10M–$100M EV",
        "boomer_density": "HIGH",
        "competition_for_deal": "LOW — real estate angle confuses most M&A bankers",
        "buyers": [
            {"name": "Lineage Logistics",        "type": "pe_rollup",   "contact": "Corp Dev",      "email": "corpdev@lineagelogistics.com",  "note": "KKR-backed, went public 2024, most aggressive"},
            {"name": "Americold Realty Trust",   "type": "public_reit", "contact": "Corp Dev",      "email": "info@americold.com",            "note": "NYSE: COLD, second largest"},
            {"name": "US Cold Storage",          "type": "pe_rollup",   "contact": "Acquisitions",  "email": "info@uscold.com",               "note": "Goldman-backed, regional builder"},
            {"name": "Nordic Logistics",         "type": "pe_rollup",   "contact": "Corp Dev",      "email": "info@nordiclogistics.com",      "note": "Smaller PE-backed rollup"},
            {"name": "VersaCold",                "type": "pe_rollup",   "contact": "M&A",           "email": "info@versacold.com",            "note": "Canada-based, expanding US"},
        ],
    },

    "medical_equipment_service": {
        "name": "Medical Equipment / Imaging Service & Repair",
        "description": (
            "Every hospital has $50M-$500M in diagnostic equipment (MRI, CT, X-ray) "
            "that must be maintained. OEM service contracts (GE, Siemens, Philips) cost "
            "3-5x what an independent service company charges. Agiliti (PE-backed) "
            "went public and is buying every regional ISP (Independent Service Provider). "
            "Most owners are former GE/Siemens biomedical engineers who went independent."
        ),
        "why_pe_loves_it": [
            "Captive hospital customers — equipment downtime costs $50K/hr",
            "Technician licensing creates 2-year barrier to entry",
            "OEM pricing 3-5x independent = massive value prop to hospitals",
            "Aging installed base = more service, not less",
            "Agiliti paying 8-12x for geographic coverage",
        ],
        "naics_codes": ["811219", "621999", "334510"],
        "ebitda_multiple": "8–12x",
        "deal_size": "$5M–$40M EV",
        "boomer_density": "HIGH",
        "competition_for_deal": "VERY LOW",
        "buyers": [
            {"name": "Agiliti Health",            "type": "pe_rollup",   "contact": "Corp Dev",      "email": "info@agilitihealth.com",        "note": "Most active buyer, KKR-backed"},
            {"name": "Sodexo / HTM Solutions",    "type": "strategic",   "contact": "Corp Dev",      "email": "info@sodexo.com",               "note": "Healthcare facilities management"},
            {"name": "Aramark Healthcare",        "type": "strategic",   "contact": "Corp Dev",      "email": "info@aramark.com",              "note": "Strategic facilities buyer"},
            {"name": "Medxcel",                   "type": "pe_rollup",   "contact": "Acquisitions",  "email": "info@medxcel.com",              "note": "Healthcare technical services"},
            {"name": "TechMed",                   "type": "pe_rollup",   "contact": "M&A",           "email": "info@techmedservices.com",      "note": "Regional ISP rollup"},
        ],
    },
}

# ═══════════════════════════════════════════════════════════════════════════════
#  NAICS → INDUSTRY LOOKUP
# ═══════════════════════════════════════════════════════════════════════════════
NAICS_MAP: Dict[str, str] = {}
for ind_key, ind in INDUSTRIES.items():
    for code in ind["naics_codes"]:
        NAICS_MAP[code] = ind_key

# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════
@dataclass
class Prospect:
    # Identity
    company_name:    str
    owner_name:      str  = ""
    city:            str  = ""
    state:           str  = ""
    address:         str  = ""
    phone:           str  = ""
    email:           str  = ""

    # Business profile
    industry_key:    str  = ""
    naics_code:      str  = ""
    employees:       int  = 0
    est_revenue:     str  = ""
    est_ebitda:      str  = ""
    business_age:    str  = ""
    ppp_loan_amt:    float = 0.0

    # Owner signals
    boomer_score:    int  = 0   # 0-100: probability owner is 60+ and wants to exit
    boomer_signals:  str  = ""

    # Sourcing
    source:          str  = "PPP Data"
    lender:          str  = ""   # their bank — useful for referral approach
    date_added:      str  = TODAY
    status:          str  = "Not Yet Contacted"
    priority:        int  = 2

    # Outreach
    owner_email_sent:    bool = False
    cpa_referral_sent:   bool = False
    notes:               str  = ""


# ═══════════════════════════════════════════════════════════════════════════════
#  BOOMER SCORING ENGINE
#  Estimates probability that owner is 60+ and considering exit
# ═══════════════════════════════════════════════════════════════════════════════
def score_boomer(row: dict) -> tuple[int, str]:
    """Return (score 0-100, signal description)."""
    score = 0
    signals = []

    # Business age (from PPP BusinessAgeDescription field)
    age_desc = row.get("BusinessAgeDescription", "").lower()
    if "existing" in age_desc or "more than 2" in age_desc:
        score += 20
        signals.append("Pre-existing business (2+ yrs)")

    # Loan amount — sweet spot is $300K-$3M (bigger = more employees = right size)
    try:
        loan = float(str(row.get("InitialApprovalAmount", "0")).replace(",", ""))
        if 300_000 <= loan <= 5_000_000:
            score += 25
            signals.append(f"PPP loan ${loan:,.0f} (sweet spot size)")
        elif 100_000 <= loan < 300_000:
            score += 10
            signals.append(f"PPP loan ${loan:,.0f} (smaller end)")
        elif loan > 5_000_000:
            score += 15
            signals.append(f"PPP loan ${loan:,.0f} (larger — may have PE already)")
    except (ValueError, TypeError):
        pass

    # Employee count — 10-150 is the target
    try:
        emp = int(str(row.get("JobsReported", "0")).replace(",", ""))
        if 10 <= emp <= 150:
            score += 20
            signals.append(f"{emp} employees (ideal size)")
        elif emp > 150:
            score += 8
            signals.append(f"{emp} employees (larger, may be too big)")
        elif 5 <= emp < 10:
            score += 5
            signals.append(f"{emp} employees (smaller end)")
    except (ValueError, TypeError):
        pass

    # Business type — Sole proprietor or S-Corp = owner-operated
    btype = row.get("BusinessType", "").lower()
    if "sole" in btype or "self-employed" in btype:
        score += 15
        signals.append("Sole proprietor / self-employed (direct owner)")
    elif "s-corp" in btype or "s_corp" in btype:
        score += 12
        signals.append("S-Corp (typical boomer structure)")
    elif "llc" in btype:
        score += 8
        signals.append("LLC (often owner-operated)")

    # Rural/suburban = less likely to have been bought already
    rural = row.get("RuralUrbanIndicator", "").upper()
    if rural == "R":
        score += 10
        signals.append("Rural market (less PE attention)")
    elif rural == "U":
        score += 0
        signals.append("Urban market")

    # Non-franchise (franchise owners are less motivated to exit independently)
    franchise = row.get("FranchiseCode", "").strip()
    if not franchise or franchise == "0":
        score += 10
        signals.append("Independent (non-franchise)")

    return min(score, 100), " | ".join(signals)


def estimate_revenue(loan_amount: float) -> tuple[str, str]:
    """Rough revenue and EBITDA estimate from PPP loan size."""
    if loan_amount <= 0:
        return "Unknown", "Unknown"
    # PPP = 2.5x monthly payroll; payroll ≈ 40% of service biz revenue
    monthly_payroll = loan_amount / 2.5
    annual_payroll = monthly_payroll * 12
    est_rev = annual_payroll / 0.40   # payroll is ~40% of revenue for services
    est_ebitda = est_rev * 0.18       # assume 18% EBITDA margin (service biz)
    return f"~${est_rev/1e6:.1f}M", f"~${est_ebitda/1e6:.1f}M"


# ═══════════════════════════════════════════════════════════════════════════════
#  PPP DATA PROCESSOR
# ═══════════════════════════════════════════════════════════════════════════════
def process_ppp(filepath: str, target_naics: Optional[List[str]] = None,
                min_loan: float = 150_000, max_loan: float = 6_000_000,
                min_employees: int = 5, max_employees: int = 200,
                limit: int = 500) -> List[Prospect]:

    prospects = []
    all_naics = target_naics or list(NAICS_MAP.keys())

    print(f"\n  Scanning PPP data: {filepath}")
    print(f"  Filters: loan ${min_loan:,.0f}–${max_loan:,.0f} | "
          f"{min_employees}–{max_employees} employees | {len(all_naics)} NAICS codes")

    with open(filepath, newline="", encoding="utf-8", errors="ignore") as f:
        reader = csv.DictReader(f)
        count = 0
        for row in reader:
            naics = str(row.get("NAICSCode", "")).strip().zfill(6)[:6]
            if naics not in all_naics:
                continue
            try:
                loan = float(str(row.get("InitialApprovalAmount", "0")).replace(",", ""))
                emp  = int(str(row.get("JobsReported", "0")).replace(",", ""))
            except (ValueError, TypeError):
                continue
            if not (min_loan <= loan <= max_loan):
                continue
            if not (min_employees <= emp <= max_employees):
                continue

            score, signals = score_boomer(row)
            rev, ebitda = estimate_revenue(loan)
            industry_key = NAICS_MAP.get(naics, "unknown")

            p = Prospect(
                company_name  = row.get("BorrowerName", "").strip().title(),
                city          = row.get("BorrowerCity", "").strip().title(),
                state         = row.get("BorrowerState", "").strip().upper(),
                address       = row.get("BorrowerAddress", "").strip().title(),
                industry_key  = industry_key,
                naics_code    = naics,
                employees     = emp,
                est_revenue   = rev,
                est_ebitda    = ebitda,
                business_age  = row.get("BusinessAgeDescription", ""),
                ppp_loan_amt  = loan,
                boomer_score  = score,
                boomer_signals= signals,
                lender        = row.get("OriginatingLender", "").strip().title(),
                source        = "PPP FOIA Data",
                priority      = 1 if score >= 70 else (2 if score >= 45 else 3),
            )
            prospects.append(p)
            count += 1
            if count >= limit:
                break

    prospects.sort(key=lambda x: (-x.boomer_score, -x.ppp_loan_amt))
    print(f"  Found {len(prospects)} qualifying prospects")
    return prospects


# ═══════════════════════════════════════════════════════════════════════════════
#  EMAIL TEMPLATE ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

def email_owner(p: Prospect) -> dict:
    """Cold outreach to business owner."""
    ind = INDUSTRIES.get(p.industry_key, {})
    ind_name = ind.get("name", "your industry")
    multiple = ind.get("ebitda_multiple", "6-9x")
    city = p.city or "your market"

    subject = f"Your {ind_name} Business — What It's Worth in Today's Market"
    body = f"""Hi,

My name is {FIRM['banker']} and I'm a mergers & acquisitions advisor at {FIRM['name']} in New York. I specialize in {ind_name} businesses in the $5M–$60M range, and I've been following the market in {city} for some time.

The market for businesses like yours is the strongest it's been in 20 years. Private equity firms and strategic acquirers are paying {multiple} EBITDA for well-run {ind_name} operations — most owners I speak with have no idea their business is worth that much.

I'm not here to rush anything. Even if you're 3-5 years from thinking about an exit, a 30-minute conversation about what your business is worth today costs you nothing and might change how you think about your next chapter.

A few things I can tell you for free:
  • Current market multiples for {ind_name} businesses your size
  • Who the most active buyers are right now (and who actually closes)
  • What makes a business worth the top of the range vs. the bottom
  • The one thing most owners do that costs them $1-3M at closing

If this resonates, I'd welcome a brief call. Complete confidentiality guaranteed — I never approach your customers, employees, or competitors without your explicit consent.

Best regards,
{FIRM['banker']}
{FIRM['title']} | {FIRM['name']}
{FIRM['email']} | {FIRM['phone']}
{FIRM['city']}"""

    return {"to": p.email or f"[find via LinkedIn: {p.company_name}]",
            "subject": subject, "body": body, "type": "Owner Outreach"}


def email_cpa(p: Prospect, cpa_name: str = "[CPA Name]",
              cpa_firm: str = "[CPA Firm]") -> dict:
    """Referral outreach to CPA who likely serves this business."""
    ind = INDUSTRIES.get(p.industry_key, {})
    ind_name = ind.get("name", "specialty services")
    multiple = ind.get("ebitda_multiple", "6-9x")

    subject = f"Referral Partnership — {ind_name} Business Sales | {FIRM['name']}"
    body = f"""Hi {cpa_name},

I'm {FIRM['banker']}, a lower middle market M&A advisor at {FIRM['name']} in New York. I focus on {ind_name} and related specialty service businesses in the $5M–$60M enterprise value range.

I'm reaching out because many of your clients in this space may be approaching the age where they're thinking about succession or a liquidity event — and most of them don't have an investment banker in their corner. That usually means they leave a significant amount on the table.

The current market is paying {multiple} EBITDA for quality {ind_name} businesses. Most owners expect 2-3x revenue. The gap between those two numbers is where your clients are losing wealth — and where a referral to the right advisor makes a material difference to their outcome.

I'd like to explore a referral relationship. When you have a client who owns a solid {ind_name} business and is thinking about the future, I handle the entire process and pay a market-rate referral fee at close (typically 1% of transaction value).

Would you be open to a 20-minute call to see if this makes sense for your clients?

Best regards,
{FIRM['banker']}
{FIRM['title']} | {FIRM['name']}
{FIRM['email']} | {FIRM['phone']}"""

    return {"to": f"[{cpa_firm}]", "subject": subject,
            "body": body, "type": "CPA Referral"}


def email_pe_buyer(buyer: dict, industry_key: str) -> dict:
    """Outreach to PE/strategic buyer with a deal."""
    ind = INDUSTRIES.get(industry_key, {})
    ind_name = ind.get("name", "specialty services")
    deal_size = ind.get("deal_size", "$5M–$40M EV")

    subject = f"[Deal Flow] {ind_name} | {deal_size} | Proprietary | {FIRM['name']}"
    body = f"""Hi {buyer.get('contact', 'Team')},

I'm {FIRM['banker']} at {FIRM['name']}, a boutique M&A advisory firm specializing in {ind_name} and related specialty services transactions.

I'm reaching out because we are actively working with owner-operators in the {ind_name} space who are exploring liquidity events — businesses that are not listed on any marketplace, not approaching other banks, and would not respond to a cold corporate development call. These are relationship-sourced, proprietary deals.

Given {buyer['name']}'s focus on this sector, I wanted to introduce myself and explore whether it makes sense to share deal flow as we develop mandates.

What I bring:
  • Proprietary access to owner-operated businesses (sourced via CPA/attorney referrals)
  • Pre-qualified sellers — motivated, realistic on price, ready to engage
  • Process management from NDA through close
  • Deep sector knowledge — I know your underwriting criteria

I'm selective about who I bring deals to. Firms that move quickly and treat sellers with respect get first looks.

Would you have 20 minutes in the next week or two?

Best regards,
{FIRM['banker']}
{FIRM['title']} | {FIRM['name']}
{FIRM['email']} | {FIRM['phone']}"""

    return {"to": buyer.get("email", ""), "subject": subject,
            "body": body, "type": "PE Buyer Outreach"}


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════════════════════
WHITE    = "FFFFFF"
NAVY     = "1F3864"
P1_GREEN = "E2EFDA"
P2_AMBER = "FFF2CC"
P3_GRAY  = "F2F2F2"
RED_SOFT = "FCE4D6"
BLUE_SFT = "D6E4F7"
ACC_GRN  = "375623"
ACC_AMB  = "7F6000"

def _f(hex_color):
    return PatternFill("solid", fgColor=hex_color)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)
def _al(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def export_excel(prospects: List[Prospect], output_path: str):
    wb = Workbook()

    # ── Sheet 1: Prospect Pipeline ──────────────────────────────────────────────
    ws = wb.active
    ws.title = "Prospect Pipeline"

    headers = [
        ("Priority",        8,  "center"),
        ("Boomer Score",   10,  "center"),
        ("Company",        30,  "left"),
        ("City / State",   18,  "left"),
        ("Industry",       28,  "left"),
        ("Est. Revenue",   12,  "center"),
        ("Est. EBITDA",    12,  "center"),
        ("Employees",      10,  "center"),
        ("PPP Loan $",     14,  "center"),
        ("Business Age",   22,  "left"),
        ("Boomer Signals", 40,  "left"),
        ("Lender (Bank)",  28,  "left"),
        ("Source",         14,  "center"),
        ("Status",         22,  "left"),
        ("Owner Name",     22,  "left"),
        ("Owner Email",    32,  "left"),
        ("Owner Phone",    16,  "center"),
        ("Notes",          36,  "left"),
    ]

    for ci, (h, w, align) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.fill      = _f(NAVY)
        cell.font      = _font(bold=True, color=WHITE, size=10)
        cell.alignment = _al(h="center", v="center")
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 26
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    def row_bg(p: Prospect) -> str:
        if p.priority == 1:   return P1_GREEN
        if p.priority == 2:   return P2_AMBER
        return P3_GRAY

    for ri, p in enumerate(prospects, 2):
        bg = row_bg(p)
        fg = ACC_GRN if p.priority == 1 else (ACC_AMB if p.priority == 2 else "444444")
        ind = INDUSTRIES.get(p.industry_key, {})

        row_data = [
            str(p.priority),
            p.boomer_score,
            p.company_name,
            f"{p.city}, {p.state}",
            ind.get("name", p.industry_key),
            p.est_revenue,
            p.est_ebitda,
            p.employees or "",
            f"${p.ppp_loan_amt:,.0f}" if p.ppp_loan_amt else "",
            p.business_age,
            p.boomer_signals,
            p.lender,
            p.source,
            p.status,
            p.owner_name,
            p.email,
            p.phone,
            p.notes,
        ]

        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = _f(bg)
            cell.font      = _font(color=fg, size=9)
            cell.border    = _border()
            cell.alignment = _al(h=headers[ci-1][2], v="top", wrap=True)

            # Priority badge
            if ci == 1:
                badge = {"1": "1B5E20", "2": "E65100", "3": "37474F"}.get(str(val), "37474F")
                cell.fill = _f(badge)
                cell.font = _font(bold=True, color=WHITE, size=10)
                cell.alignment = _al(h="center", v="center")

            # Boomer score color
            elif ci == 2:
                try:
                    s = int(val)
                    if s >= 70:   cell.fill = _f("C6EFCE")
                    elif s >= 45: cell.fill = _f("FFEB9C")
                    else:         cell.fill = _f("FFC7CE")
                except (ValueError, TypeError):
                    pass
                cell.font = _font(bold=(int(val) >= 70 if val else False), size=9)

        ws.row_dimensions[ri].height = 40

    # ── Sheet 2: Buyer Database ─────────────────────────────────────────────────
    wb2 = wb.create_sheet("PE Buyer Database")
    buyer_hdrs = [
        ("Industry",         28, "left"),
        ("Multiple",         12, "center"),
        ("Deal Size",        16, "center"),
        ("Buyer Name",       30, "left"),
        ("Type",             16, "center"),
        ("Contact",          24, "left"),
        ("Email",            36, "left"),
        ("Notes",            50, "left"),
    ]
    for ci, (h, w, align) in enumerate(buyer_hdrs, 1):
        cell = wb2.cell(row=1, column=ci, value=h)
        cell.fill = _f(NAVY); cell.font = _font(bold=True, color=WHITE, size=10)
        cell.alignment = _al(h="center"); cell.border = _border()
        wb2.column_dimensions[get_column_letter(ci)].width = w
    wb2.row_dimensions[1].height = 26
    wb2.freeze_panes = "A2"
    wb2.auto_filter.ref = "A1:H1"

    row_idx = 2
    colors = ["D6E4F7", "E2EFDA", "FFF2CC", "FCE4D6", "F2F2F2",
              "E8D5F5", "D4EDDA", "FFE0B2", "E3F2FD", "FBE9E7"]
    for i, (key, ind) in enumerate(INDUSTRIES.items()):
        bg = colors[i % len(colors)]
        for buyer in ind["buyers"]:
            row_data = [
                ind["name"],
                ind["ebitda_multiple"],
                ind["deal_size"],
                buyer["name"],
                buyer["type"].replace("_", " ").title(),
                buyer["contact"],
                buyer["email"],
                buyer["note"],
            ]
            for ci, val in enumerate(row_data, 1):
                cell = wb2.cell(row=row_idx, column=ci, value=val)
                cell.fill = _f(bg); cell.border = _border()
                cell.alignment = _al(h=buyer_hdrs[ci-1][2], v="center", wrap=True)
                cell.font = _font(size=9)
                if ci == 7:  # email
                    cell.font = Font(color="1155CC", underline="single", size=9, name="Calibri")
            wb2.row_dimensions[row_idx].height = 20
            row_idx += 1

    # ── Sheet 3: Email Templates ────────────────────────────────────────────────
    wb3 = wb.create_sheet("Email Templates")
    wb3.column_dimensions["A"].width = 22
    wb3.column_dimensions["B"].width = 80

    tmpl_hdrs = [("Template Type", NAVY), ("Content", NAVY)]
    for ci, (h, bg) in enumerate(tmpl_hdrs, 1):
        cell = wb3.cell(row=1, column=ci, value=h)
        cell.fill = _f(bg); cell.font = _font(bold=True, color=WHITE, size=10)
        cell.alignment = _al(h="center"); cell.border = _border()
    wb3.freeze_panes = "A2"

    # Generate sample templates for first industry
    sample_p = Prospect(company_name="Smith Fire Protection",
                        city="Columbus", state="OH",
                        industry_key="fire_safety", employees=18,
                        est_revenue="~$4.2M", est_ebitda="~$756K",
                        ppp_loan_amt=350000, boomer_score=82)
    sample_buyer = INDUSTRIES["fire_safety"]["buyers"][0]

    templates = [
        ("1. OWNER OUTREACH\n(cold email to business owner)", email_owner(sample_p)["body"]),
        ("2. CPA REFERRAL\n(to their accountant)", email_cpa(sample_p)["body"]),
        ("3. PE BUYER OUTREACH\n(build buy-side relationships)", email_pe_buyer(sample_buyer, "fire_safety")["body"]),
        ("4. FOLLOW-UP #1\n(1 week after initial outreach)",
         f"Hi [Name],\n\nI wanted to follow up on my note from last week regarding the M&A market for [industry] businesses.\n\nI'll keep this brief — the market is very active right now and I'm working with a handful of buyers actively looking for businesses like yours. If the timing isn't right, no problem at all. But if you've thought about the future of your business even once in the last year, it's worth 20 minutes.\n\nWould any time this week work?\n\n{FIRM['banker']}\n{FIRM['email']} | {FIRM['phone']}"),
        ("5. LENDER REFERRAL\n(to the business's originating bank)",
         f"Hi [Loan Officer Name],\n\nI'm {FIRM['banker']}, an M&A advisor at {FIRM['name']} in New York. I work with owner-operated businesses in specialty services — HVAC, fire safety, calibration labs, specialty recycling, and similar sectors.\n\nI'm reaching out because your institution has relationships with business owners who are approaching retirement age and don't have a succession plan. Many of them will eventually sell — the question is whether they get the right outcome.\n\nI'd love to explore a referral arrangement. When you have a commercial banking client thinking about their exit, I handle the entire M&A process and can pay a referral fee at close.\n\nWould a brief call make sense?\n\n{FIRM['banker']}\n{FIRM['email']} | {FIRM['phone']}"),
    ]

    row_idx = 2
    tmpl_colors = [P1_GREEN, P2_AMBER, BLUE_SFT, "FCE4D6", "E8D5F5"]
    for i, (label, content) in enumerate(templates):
        cell_a = wb3.cell(row=row_idx, column=1, value=label)
        cell_b = wb3.cell(row=row_idx, column=2, value=content)
        bg = tmpl_colors[i % len(tmpl_colors)]
        for cell in (cell_a, cell_b):
            cell.fill = _f(bg); cell.border = _border()
            cell.alignment = _al(v="top", wrap=True)
            cell.font = _font(size=9)
        cell_a.font = _font(bold=True, size=9, color=NAVY)
        wb3.row_dimensions[row_idx].height = max(120, len(content) // 3)
        row_idx += 1

    # ── Sheet 4: Strategy Brief ─────────────────────────────────────────────────
    wb4 = wb.create_sheet("Strategy Brief")
    wb4.column_dimensions["A"].width = 30
    wb4.column_dimensions["B"].width = 70

    strategy = [
        ("THE SILVER TSUNAMI PLAYBOOK", "",                    NAVY,    WHITE,  True,  13),
        ("The Opportunity",  "10,000 boomers turn 65 every day. $10T in private business "
                             "wealth will transfer over the next 15 years. Most owners have "
                             "a CPA but no banker. You are the bridge.",
                             P1_GREEN, ACC_GRN, False, 10),
        ("", "", WHITE, "000000", False, 10),
        ("WHY NICHE?", "",  NAVY, WHITE, True, 11),
        ("Avoid the obvious", "HVAC, plumbing, pest control — every broker and junior banker "
                              "in the country is calling them. You need categories where "
                              "no one is calling. Fire safety, elevator inspection, calibration "
                              "labs, specialty recycling — the owner has never spoken to a banker.",
                              P2_AMBER, ACC_AMB, False, 10),
        ("", "", WHITE, "000000", False, 10),
        ("THE PPP DATA EDGE", "", NAVY, WHITE, True, 11),
        ("What it is",       "The SBA released complete PPP loan records — 5.2M businesses, "
                             "all public. Download from data.sba.gov/dataset/ppp-foia",
                             BLUE_SFT, "1F3864", False, 10),
        ("What you get",     "Business name, address, NAICS code, loan amount (proxy for "
                             "payroll/revenue), employee count, their bank, and business age. "
                             "Everything you need to build a prospect list in any industry.",
                             BLUE_SFT, "1F3864", False, 10),
        ("The play",         "Filter by NAICS code for your target industries. Filter loan "
                             "$300K-$5M (= ~$9M-$60M revenue). Their lender is your referral "
                             "contact. The bank officer knows these clients are aging.",
                             BLUE_SFT, "1F3864", False, 10),
        ("", "", WHITE, "000000", False, 10),
        ("FINDING OWNER AGE", "", NAVY, WHITE, True, 11),
        ("LinkedIn method",  "Search owner name + company. Check education dates. "
                             "Undergrad 1975-1985 = born 1953-1963 = age 61-73. High priority.",
                             P1_GREEN, ACC_GRN, False, 10),
        ("Incorporation date","Secretary of State filings are public in most states. "
                              "Company founded 1985-2000 + owner-operated = very likely 60+.",
                              P1_GREEN, ACC_GRN, False, 10),
        ("Google method",    "'[Company name] founded' or '[Owner name] [city]' — local "
                             "news articles often mention 'started his business 30 years ago'.",
                             P1_GREEN, ACC_GRN, False, 10),
        ("The lender angle", "PPP data shows their originating bank. Call the loan officer: "
                             "'I work with specialty service businesses in your area thinking "
                             "about succession planning. Who on your commercial portfolio is "
                             "approaching retirement?' They will tell you.",
                             P2_AMBER, ACC_AMB, False, 10),
        ("", "", WHITE, "000000", False, 10),
        ("90-DAY LAUNCH PLAN", "", NAVY, WHITE, True, 11),
        ("Week 1-2",  "Pick 2 verticals. Fire safety + calibration recommended. "
                      "Map all active PE buyers. Get the M&A contact on the phone — "
                      "tell them you're sourcing in their space. They WILL take your call.",
                      P1_GREEN, ACC_GRN, True, 10),
        ("Week 3-4",  "Download PPP data for your target states. Run this script. "
                      "Build a list of 300+ target companies. Enrich with LinkedIn.",
                      P2_AMBER, ACC_AMB, True, 10),
        ("Week 5-8",  "Build CPA referral network. Target 50 regional accounting firms "
                      "in secondary markets (Ohio, Texas, Southeast, not NYC). "
                      "Referral fee: 1% of deal value at close.",
                      P1_GREEN, ACC_GRN, True, 10),
        ("Week 9-12", "Direct owner outreach. 200 emails, 5-10% response rate, "
                      "expect 2-3 real conversations per 100. One signed mandate "
                      "changes your whole pipeline.",
                      P2_AMBER, ACC_AMB, True, 10),
        ("Month 4+",  "First mandate = case study = credibility = referrals. "
                      "Do NOT take on every deal. Only mandates where you know "
                      "3+ active buyers who will move fast.",
                      P1_GREEN, ACC_GRN, True, 10),
        ("", "", WHITE, "000000", False, 10),
        ("THE MATH",  "One $25M fire safety deal at 4% fee = $1M. "
                      "Do 4 deals/year = $4M in revenue with 2 bankers. "
                      "You don't need 50 deals. You need 4 great ones.",
                      "C6EFCE", ACC_GRN, True, 11),
    ]

    row_idx = 1
    for label, content, bg, fg, bold, sz in strategy:
        ca = wb4.cell(row=row_idx, column=1, value=label)
        cb = wb4.cell(row=row_idx, column=2, value=content)
        for c in (ca, cb):
            c.fill = _f(bg); c.border = _border()
            c.alignment = _al(v="top", wrap=True)
            c.font = _font(bold=bold, color=fg, size=sz)
        wb4.row_dimensions[row_idx].height = max(20, len(content)//4) if content else 18
        row_idx += 1

    wb.save(output_path)
    print(f"\n  ✓ Saved → {output_path}")
    print(f"  Sheets: Prospect Pipeline | PE Buyer Database | Email Templates | Strategy Brief")


# ═══════════════════════════════════════════════════════════════════════════════
#  SUMMARY PRINT
# ═══════════════════════════════════════════════════════════════════════════════
def print_summary(prospects: List[Prospect]):
    p1 = [p for p in prospects if p.priority == 1]
    p2 = [p for p in prospects if p.priority == 2]
    p3 = [p for p in prospects if p.priority == 3]

    print("\n" + "═"*72)
    print("  SILVER TSUNAMI — DEAL SOURCING ENGINE")
    print("═"*72)
    print(f"  Total prospects found:  {len(prospects)}")
    print(f"  Priority 1 (score 70+): {len(p1)}  ← call these first")
    print(f"  Priority 2 (score 45+): {len(p2)}")
    print(f"  Priority 3 (score <45): {len(p3)}")
    print(f"\n  Industries covered:     {len(INDUSTRIES)}")
    print(f"  PE buyers in database:  {sum(len(v['buyers']) for v in INDUSTRIES.values())}")
    print("─"*72)

    if p1:
        print(f"\n  TOP 10 HIGHEST SCORING PROSPECTS\n")
        for p in p1[:10]:
            ind = INDUSTRIES.get(p.industry_key, {})
            print(f"  [{p.boomer_score:3d}] {p.company_name:<35} {p.city}, {p.state}")
            print(f"       {ind.get('name',''):<30} | Rev: {p.est_revenue} | EBITDA: {p.est_ebitda}")
            print(f"       Bank: {p.lender[:40]}")
            print()

    print("─"*72)
    print(f"\n  INDUSTRY BREAKDOWN\n")
    by_industry: Dict[str, List[Prospect]] = {}
    for p in prospects:
        by_industry.setdefault(p.industry_key, []).append(p)
    for key, ps in sorted(by_industry.items(), key=lambda x: -len(x[1])):
        ind = INDUSTRIES.get(key, {})
        avg_score = sum(p.boomer_score for p in ps) // len(ps) if ps else 0
        print(f"  {ind.get('name', key):<38} {len(ps):3d} prospects | avg score {avg_score}")
    print("═"*72)


# ═══════════════════════════════════════════════════════════════════════════════
#  MANUAL PROSPECT BUILDER (no PPP file needed)
# ═══════════════════════════════════════════════════════════════════════════════
SAMPLE_MANUAL: List[Prospect] = [
    Prospect("Acme Elevator Services",     "Bob Kowalski",  "Cleveland",   "OH",
             industry_key="elevator",         employees=22, est_revenue="~$4.8M",
             est_ebitda="~$864K", boomer_score=85, priority=1,
             boomer_signals="LinkedIn: graduated 1978 | Company founded 1989 | S-Corp",
             lender="First National Bank of Ohio", source="LinkedIn Research",
             notes="Son works at company but has no interest in owning it. Referred by CPA."),

    Prospect("Tri-State Fire Protection",  "Gary Malone",   "Pittsburgh",  "PA",
             industry_key="fire_safety",       employees=34, est_revenue="~$7.2M",
             est_ebitda="~$1.3M", boomer_score=91, priority=1,
             boomer_signals="Company founded 1987 | State inspection license since 1988 | Sole prop",
             lender="PNC Bank", source="PPP Data + LinkedIn",
             notes="Gary mentioned retirement in a 2022 local news article. No kids in business."),

    Prospect("Precision Calibration Labs",  "Susan Whitfield","Indianapolis","IN",
             industry_key="calibration",       employees=12, est_revenue="~$2.8M",
             est_ebitda="~$560K", boomer_score=77, priority=1,
             boomer_signals="ISO 17025 since 1995 | LinkedIn grad 1981 | No succession",
             lender="Regions Bank", source="Industry Association Directory",
             notes="Serves Eli Lilly and local aerospace. Transcat would pay 10-12x for this."),

    Prospect("Green Path E-Waste",          "Tom Cipriani",  "Phoenix",     "AZ",
             industry_key="specialty_recycling", employees=28, est_revenue="~$5.1M",
             est_ebitda="~$918K", boomer_score=72, priority=1,
             boomer_signals="EPA license since 1996 | PPP loan $612K | S-Corp",
             lender="Wells Fargo", source="PPP Data",
             notes="Battery recycling + e-waste. Clean Earth would pay up for AZ license."),

    Prospect("Midwest Cold Storage",        "Frank Reardon", "St. Louis",   "MO",
             industry_key="cold_storage",      employees=41, est_revenue="~$8.9M",
             est_ebitda="~$1.6M", boomer_score=68, priority=2,
             boomer_signals="Facility built 1991 | LinkedIn grad 1979 | Owns the building",
             lender="Commerce Bank", source="SBA 7a Loan Data",
             notes="Owns 65,000 sq ft freezer. Real estate play on top of operating business."),
]


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Silver Tsunami — Boomer Business Sourcing Engine",
        formatter_class=argparse.RawTextHelpFormatter,
    )
    parser.add_argument("--ppp",      metavar="FILE",   help="Path to SBA PPP CSV (download from data.sba.gov)")
    parser.add_argument("--naics",    metavar="SECTOR", help="Industry key to filter: " + " | ".join(INDUSTRIES.keys()))
    parser.add_argument("--state",    metavar="STATE",  help="Filter by state abbreviation (e.g. OH, TX)")
    parser.add_argument("--min-loan", type=float, default=150_000, help="Min PPP loan $ (default 150K)")
    parser.add_argument("--max-loan", type=float, default=6_000_000, help="Max PPP loan $ (default 6M)")
    parser.add_argument("--limit",    type=int,   default=500,   help="Max prospects to return (default 500)")
    parser.add_argument("--manual",   action="store_true",        help="Load sample manual prospects instead of PPP data")
    parser.add_argument("--output",   metavar="FILE", default="Silver_Tsunami_Pipeline.xlsx", help="Output Excel file")
    args = parser.parse_args()

    if args.ppp:
        target_naics = None
        if args.naics:
            key = args.naics.lower()
            if key not in INDUSTRIES:
                print(f"  Unknown industry: {key}. Options: {list(INDUSTRIES.keys())}")
                sys.exit(1)
            target_naics = INDUSTRIES[key]["naics_codes"]

        prospects = process_ppp(
            args.ppp,
            target_naics=target_naics,
            min_loan=args.min_loan,
            max_loan=args.max_loan,
            limit=args.limit,
        )
        if args.state:
            prospects = [p for p in prospects if p.state == args.state.upper()]
            print(f"  Filtered to {args.state.upper()}: {len(prospects)} prospects")

    elif args.manual:
        prospects = SAMPLE_MANUAL
        print(f"\n  Loaded {len(prospects)} sample manual prospects")

    else:
        # No PPP file: show the buyer database + templates + strategy only
        print("\n  No PPP file specified — generating buyer database + templates only.")
        print("  Tip: download PPP data from https://data.sba.gov/dataset/ppp-foia")
        print("  Then run: python silver_tsunami.py --ppp your_state.csv\n")
        prospects = SAMPLE_MANUAL

    print_summary(prospects)
    export_excel(prospects, args.output)

    print(f"\n  HOW TO GET PPP DATA:")
    print(f"  1. Go to https://data.sba.gov/dataset/ppp-foia")
    print(f"  2. Download state CSV files (e.g. 'public_150k_plus.csv' for larger loans)")
    print(f"  3. Run: python silver_tsunami.py --ppp public_150k_plus.csv --naics fire_safety")
    print(f"\n  INDUSTRY KEYS: {' | '.join(INDUSTRIES.keys())}\n")


if __name__ == "__main__":
    main()
