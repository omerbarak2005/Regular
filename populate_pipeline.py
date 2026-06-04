#!/usr/bin/env python3
"""
M&A Origination Pipeline Builder
Populates Jim's workbook Pipeline tab with researched seed target contacts.
Usage: python populate_pipeline.py [--output MA_Pipeline_Populated.xlsx]
"""

import argparse
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TODAY = date.today().strftime("%Y-%m-%d")

# ── Colors ──────────────────────────────────────────────────────────────────
NAV      = "1F3864"
WHITE    = "FFFFFF"
GREEN_BG = "E2EFDA"
GREEN_FG = "375623"
AMBER_BG = "FFF2CC"
AMBER_FG = "7F6000"
RED_BG   = "FCE4D6"
RED_FG   = "843C0C"
GRAY_BG  = "F2F2F2"
GRAY_FG  = "595959"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")
def _border():
    s = Side(style="thin", color="BDD7EE")
    return Border(left=s, right=s, top=s, bottom=s)
def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


# ── Pipeline columns (matches Jim's workbook exactly) ───────────────────────
PIPELINE_COLS = [
    ("Company Name",       22, False, "left"),
    ("Website",            26, False, "left"),
    ("State",               6, False, "center"),
    ("City/HQ",            16, False, "left"),
    ("Industry",           20, False, "left"),
    ("Est. Revenue",       14, False, "right"),
    ("Est. EBITDA",        13, False, "right"),
    ("Est. Employees",     14, False, "center"),
    ("Ownership Type",     16, False, "left"),
    ("Year Founded",       12, False, "center"),
    ("Multi-Location?",    14, False, "center"),
    ("Priority Score",     13, False, "center"),
    ("Status",             18, False, "left"),
    ("Contact Name",       20, False, "left"),
    ("Title",              22, False, "left"),
    ("Email",              36, False, "left"),
    ("Phone",              16, False, "center"),
    ("LinkedIn URL",       40, False, "left"),
    ("Last Outreach Date", 16, False, "center"),
    ("Next Step",          30, True,  "left"),
    ("Notes / Why Attractive", 46, True, "left"),
    ("Verification Needed",    24, True,  "left"),
]

# ── Researched pipeline rows ─────────────────────────────────────────────────
PIPELINE_ROWS = [
    # ── PRIORITY 1 — send outreach this week ──────────────────────────────
    {
        "Company Name":       "Aiello Home Services",
        "Website":            "aiellohomeservices.com",
        "State":              "CT",
        "City/HQ":            "Windsor Locks",
        "Industry":           "HVAC / Plumbing / Electrical",
        "Est. Revenue":       "$22M",
        "Est. EBITDA":        "$3M–$4M",
        "Est. Employees":     "60",
        "Ownership Type":     "Owner-operated (Jezouit purchased 1980)",
        "Year Founded":       "~1935 (Jezouit acquired 1980)",
        "Multi-Location?":    "No",
        "Priority Score":     "1",
        "Status":             "Researching",
        "Contact Name":       "Michael Jezouit",
        "Title":              "President & CEO",
        "Email":              "m.jezouit@aiellohomeservices.com ✓",
        "Phone":              "(860) 292-2600",
        "LinkedIn URL":       "linkedin.com/in/michael-jezouit-0b18a0126/",
        "Last Outreach Date": "",
        "Next Step":          "SEND NOW — email confirmed. Reference 90-yr legacy + succession planning.",
        "Notes / Why Attractive": "Michael Jezouit purchased Aiello Plumbing & Heating in 1980 and grew it to $22M revenue and 60+ employees — he IS the owner, not a hired CEO. Classic owner-operator who built it himself = primary decision-maker on any exit. $22M revenue confirmed (D&B 2025). EMAIL CONFIRMED via ZoomInfo + RocketReach.",
        "Verification Needed": "Confirm no PE minority investment; check if any roll-up (Apex, HomeX) has already approached; confirm Michael's exit timeline",
    },
    {
        "Company Name":       "Mazza Recycling Services",
        "Website":            "mazzarecycling.com",
        "State":              "NJ",
        "City/HQ":            "Tinton Falls",
        "Industry":           "Recycling / Waste Management",
        "Est. Revenue":       "$25M–$40M",
        "Est. EBITDA":        "$3M–$5M",
        "Est. Employees":     "100–175",
        "Ownership Type":     "Family (2nd gen)",
        "Year Founded":       "1964",
        "Multi-Location?":    "No",
        "Priority Score":     "1",
        "Status":             "Researching",
        "Contact Name":       "James Mazza Jr.",
        "Title":              "President & CEO",
        "Email":              "j.mazza@mazzarecycling.com",
        "Phone":              "(732) 992-8055",
        "LinkedIn URL":       "linkedin.com/in/james-mazza-7614b316b/",
        "Last Outreach Date": "",
        "Next Step":          "Email + LinkedIn. Central NJ recycler. Reference ESG and PE appetite for recycling.",
        "Notes / Why Attractive": "60-year-old family recycling company. Dominick and James Mazza Sr. founded 1964; James Jr. now runs it. $25-40M revenue. Central NJ recycling + C&D waste — PE loves recurring, route-based environmental services. Recently acquired Liberty Waste (2024), showing growth appetite.",
        "Verification Needed": "Confirm no PE involvement; check post-Liberty acquisition financials; verify email format",
    },
    {
        "Company Name":       "Mowrey Elevator",
        "Website":            "mowreyelevator.com",
        "State":              "FL",
        "City/HQ":            "Marianna",
        "Industry":           "Elevator Sales, Service & Modernization",
        "Est. Revenue":       "$20M–$50M",
        "Est. EBITDA":        "$3M–$8M",
        "Est. Employees":     "100–400",
        "Ownership Type":     "Family (founder-led)",
        "Year Founded":       "1976",
        "Multi-Location?":    "Yes (Marianna 380K sq ft plant + multiple SE service centers)",
        "Priority Score":     "1",
        "Status":             "Researching",
        "Contact Name":       "Tim Mowrey Sr.",
        "Title":              "Founder & President",
        "Email":              "t.mowrey@mowreyelevator.com",
        "Phone":              "(850) 526-4111",
        "LinkedIn URL":       "linkedin.com/in/tim-mowrey-29a39342/",
        "Last Outreach Date": "",
        "Next Step":          "Call + email Tim directly. Reference elevator roll-up activity (Arcline, Kone, etc).",
        "Notes / Why Attractive": "Founded 1976, 15,000+ elevators installed, 380K sq ft manufacturing plant in Marianna FL. Tim Mowrey Sr. known as 'The Elevator Man.' Company LARGER than originally estimated — revenue likely $20M-$50M range (hundreds of employees per website). Family-owned Southeast regional leader in elevator manufacturing + service. APi Group and 3Phase both buying elevator companies.",
        "Verification Needed": "Verify actual revenue before approaching — may be above typical LMM threshold; confirm Tim Sr. still primary decision-maker; confirm no PE approach",
    },
    {
        "Company Name":       "Russell Reid Waste Management",
        "Website":            "russellreid.com",
        "State":              "NJ",
        "City/HQ":            "Edison",
        "Industry":           "Septic / Liquid Waste",
        "Est. Revenue":       "$30M–$60M",
        "Est. EBITDA":        "$5M–$10M",
        "Est. Employees":     "100–200",
        "Ownership Type":     "Family (2nd gen)",
        "Year Founded":       "1943",
        "Multi-Location?":    "Yes (5 locations)",
        "Priority Score":     "1",
        "Status":             "Researching",
        "Contact Name":       "Gary M. Weiner",
        "Title":              "President",
        "Email":              "gweiner@russellreid.com",
        "Phone":              "(800) 356-4468",
        "LinkedIn URL":       "linkedin.com/company/russellreidresponsiblewastemanagement",
        "Last Outreach Date": "",
        "Next Step":          "Verify Gary Weiner email; call main line. NJ's largest independent liquid waster.",
        "Notes / Why Attractive": "80-year family business, 5 NJ/NY/PA/DE locations. Morton Weiner acquired 1943 company in 1981; son Gary now President. Septic+grease trap = recurring, non-discretionary. Scale + geography = ideal PE platform. Largest independent liquid waste in NJ.",
        "Verification Needed": "Confirm Gary Weiner email format; confirm no active sale process; verify revenue",
    },

    # ── PRIORITY 2 — verify, then send ────────────────────────────────────
    {
        "Company Name":       "Burkholder's HVAC",
        "Website":            "burkholders-hvac.com",
        "State":              "PA",
        "City/HQ":            "Emmaus",
        "Industry":           "HVAC",
        "Est. Revenue":       "$8M–$15M",
        "Est. EBITDA":        "$1M–$2.5M",
        "Est. Employees":     "40–70",
        "Ownership Type":     "Family (2nd gen)",
        "Year Founded":       "1960",
        "Multi-Location?":    "No",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "Robert (Bob) Burkholder",
        "Title":              "President",
        "Email":              "r.burkholder@burkholders-hvac.com",
        "Phone":              "(610) 816-6889",
        "LinkedIn URL":       "burkholders-hvac.com/our-team/",
        "Last Outreach Date": "",
        "Next Step":          "Call main line; confirm Bob's email. 65-year Lehigh Valley HVAC — strong succession signal.",
        "Notes / Why Attractive": "Founded 1960 by Carl Burkholder; son Bob took over in 2000. 65+ years, Lehigh Valley PA. Carrier Factory Authorized Dealer = quality brand. Recently acquired Lande Heating (Bethlehem) — growth appetite. Bob is likely looking toward an exit in the next 3-5 years.",
        "Verification Needed": "Confirm Bob's email (try bob@burkholders-hvac.com); confirm size; check if any PE approached",
    },
    {
        "Company Name":       "Confires Fire Protection",
        "Website":            "confires.com",
        "State":              "NJ",
        "City/HQ":            "South Plainfield",
        "Industry":           "Fire Protection / Life Safety",
        "Est. Revenue":       "$5M–$10M",
        "Est. EBITDA":        "$750K–$1.5M",
        "Est. Employees":     "33",
        "Ownership Type":     "Privately held",
        "Year Founded":       "~1985",
        "Multi-Location?":    "Yes (affiliated offices)",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "William Wrublevski",
        "Title":              "Owner",
        "Email":              "william.wrublevski@confires.com",
        "Phone":              "(908) 822-2700",
        "LinkedIn URL":       "linkedin.com/company/confires-fire-protection-services-llc/",
        "Last Outreach Date": "",
        "Next Step":          "Call (908) 822-2700; email William. 40+ years NJ/DE/PA fire protection. Pye-Barker did 57 deals in 2024 — obvious buyer here.",
        "Notes / Why Attractive": "40+ years NJ/DE/PA fire protection. Recurring inspection/service revenue. William Wrublevski (owner), Mark Calleo (VP). Email format first.last@confires.com CONFIRMED. Fire protection = PE's current favorite roll-up (Pye-Barker, Fortis dominating).",
        "Verification Needed": "Confirm Wrublevski ownership stake; confirm still independent; check size",
    },
    {
        "Company Name":       "T.F. O'Brien & Co.",
        "Website":            "tfobrien.com",
        "State":              "NY",
        "City/HQ":            "New Hyde Park",
        "Industry":           "HVAC / Cooling & Heating",
        "Est. Revenue":       "$10M–$18M",
        "Est. EBITDA":        "$1.5M–$2.5M",
        "Est. Employees":     "25–35",
        "Ownership Type":     "Family (3rd gen)",
        "Year Founded":       "1934",
        "Multi-Location?":    "No",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "Kerry O'Brien",
        "Title":              "Owner",
        "Email":              "k.obrien@tfobrien.com",
        "Phone":              "(516) 488-1800",
        "LinkedIn URL":       "linkedin.com/in/kerry-o-brien-b7906b6/",
        "Last Outreach Date": "",
        "Next Step":          "LinkedIn + email. 90+ yr Long Island HVAC. Multiple O'Brien siblings = succession complexity.",
        "Notes / Why Attractive": "90+ year Long Island HVAC institution (1934). 3rd gen family (Kerry, Thomas Jr., Christopher O'Brien). Multiple heirs = succession complexity = exit motivation. Strong brand in high-income Long Island market.",
        "Verification Needed": "Confirm current co-owner names; confirm revenue; verify email format",
    },
    {
        "Company Name":       "Suburban Disposal",
        "Website":            "suburbandisposal.com",
        "State":              "NJ",
        "City/HQ":            "Fairfield",
        "Industry":           "Waste Management / Hauling",
        "Est. Revenue":       "$17M",
        "Est. EBITDA":        "$2.5M–$4M",
        "Est. Employees":     "~100",
        "Ownership Type":     "Family",
        "Year Founded":       "1979",
        "Multi-Location?":    "No",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "Chris Roselle",
        "Title":              "President",
        "Email":              "c.roselle@suburbandisposalinc.com",
        "Phone":              "(973) 227-7020",
        "LinkedIn URL":       "linkedin.com/in/john-roselle-645399a0/",
        "Last Outreach Date": "",
        "Next Step":          "Call main line (973) 227-7020; ask for Chris Roselle (President). John Roselle is a Partner. $17M revenue confirmed, ~100 employees, largest municipal hauler in NJ.",
        "Notes / Why Attractive": "45-year Northern NJ waste hauler, $17M revenue confirmed, ~100 employees (D&B). Roselle family: Chris Roselle (President), Anthony Roselle (Treasurer), John Roselle (Partner). Self-described largest municipal hauler in NJ — significant market position. Fairfield NJ location strategic for a route-based platform acquirer.",
        "Verification Needed": "Confirm who the owner is (John Roselle title unclear); confirm not PE-backed",
    },
    {
        "Company Name":       "A Royal Flush",
        "Website":            "aroyalflush.com",
        "State":              "CT",
        "City/HQ":            "Bridgeport",
        "Industry":           "Portable Sanitation",
        "Est. Revenue":       "$15M–$25M",
        "Est. EBITDA":        "$2M–$4M",
        "Est. Employees":     "75–125",
        "Ownership Type":     "Privately held",
        "Year Founded":       "1992",
        "Multi-Location?":    "Yes (CT/NY/NJ/PA/MA/SC/FL)",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "Tim Butler",
        "Title":              "Principal Owner & Chairman",
        "Email":              "tbutler@aroyalflush.com",
        "Phone":              "(203) 333-1400",
        "LinkedIn URL":       "linkedin.com/company/a-royal-flush-inc/",
        "Last Outreach Date": "",
        "Next Step":          "Email Tim Butler. 10,000+ unit fleet, 7-state footprint = platform-level portable sanitation.",
        "Notes / Why Attractive": "30+ year CT portable sanitation leader. 10,000+ toilets, 80 trailers, 90 trucks. 7-state footprint (CT/NY/NJ/PA/MA/SC/FL) = true regional platform. Tim Butler took over after founder Bill Malone passed 2017. COO Mauro DaSilva. Portable sanitation = route-based, recurring, PE-loved.",
        "Verification Needed": "Confirm Tim's email format; confirm ownership structure; check if any PE contacted",
    },
    {
        "Company Name":       "Texas Outhouse",
        "Website":            "texasouthouse.com",
        "State":              "TX",
        "City/HQ":            "Houston",
        "Industry":           "Portable Sanitation",
        "Est. Revenue":       "$10M–$20M",
        "Est. EBITDA":        "$1.5M–$3M",
        "Est. Employees":     "50–100",
        "Ownership Type":     "Family (Carl family)",
        "Year Founded":       "1999",
        "Multi-Location?":    "Yes",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "Paul Carl",
        "Title":              "President",
        "Email":              "paul.carl@texasouthouse.com",
        "Phone":              "(713) 690-9800",
        "LinkedIn URL":       "linkedin.com/company/texas-outhouse/",
        "Last Outreach Date": "",
        "Next Step":          "Email Paul Carl. Sister company Gainsborough Waste = potential double deal.",
        "Notes / Why Attractive": "Houston metro portable sanitation leader. Carl family: W. Noble Carl III (father), Paul Carl (President/VP), Noble Jr. (brother). Also runs sister company Gainsborough Waste. Family-operated since 1999. Houston market = large construction base = strong recurring revenue.",
        "Verification Needed": "Confirm Paul's email; confirm not PE-approached; check Gainsborough Waste ownership separately",
    },
    {
        "Company Name":       "Revolution Recovery",
        "Website":            "revolutionrecovery.com",
        "State":              "PA",
        "City/HQ":            "Philadelphia",
        "Industry":           "Recycling / C&D Waste",
        "Est. Revenue":       "$30M–$60M",
        "Est. EBITDA":        "$4M–$8M",
        "Est. Employees":     "100–200",
        "Ownership Type":     "Founder-owned",
        "Year Founded":       "2004",
        "Multi-Location?":    "Yes (Phila + Wilmington DE + Allentown PA)",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "Jonathan Wybar",
        "Title":              "Founder & Owner",
        "Email":              "jwybar@revolutionrecovery.com",
        "Phone":              "(215) 333-6505",
        "LinkedIn URL":       "linkedin.com/company/revolution-recovery/",
        "Last Outreach Date": "",
        "Next Step":          "Email Jonathan + call. C&D recycling = strong PE interest. 3 locations = platform.",
        "Notes / Why Attractive": "Jon Wybar co-founded 2004 with Avi Golen; Wybar is now primary owner. Processing 600 tons/day at Philadelphia facility alone + Delaware + Allentown locations. Revenue likely $30M-$60M range — much larger than originally estimated. C&D recycling is fragmented and PE-active. No PE backing confirmed. Recently bought a Superfund site, showing capital appetite. Confirm Avi Golen's current stake.",
        "Verification Needed": "Confirm Jonathan's email; confirm not PE-backed; verify revenue/employee count",
    },
    {
        "Company Name":       "Bosworth Air Conditioning",
        "Website":            "bosworthac.com",
        "State":              "TX",
        "City/HQ":            "Galveston",
        "Industry":           "HVAC",
        "Est. Revenue":       "$8M–$15M",
        "Est. EBITDA":        "$1M–$2M",
        "Est. Employees":     "40–70",
        "Ownership Type":     "Family (2nd gen)",
        "Year Founded":       "1959",
        "Multi-Location?":    "No",
        "Priority Score":     "3",
        "Status":             "Researching",
        "Contact Name":       "Jerry Bosworth Jr.",
        "Title":              "Owner",
        "Email":              "jerry@bosworthac.com",
        "Phone":              "(409) 762-0418",
        "LinkedIn URL":       "acca.org/board/jerry-bosworth/",
        "Last Outreach Date": "",
        "Next Step":          "Verify if Jerry Sr. or Jr. runs it now. 65+ year Gulf Coast HVAC.",
        "Notes / Why Attractive": "65-year Galveston TX HVAC family business. Jerry Bosworth Sr. founded 1959; Jerry Jr. is current owner. Gulf Coast market = year-round AC demand. ACCA board member = industry credibility. Smaller size but solid family succession candidate.",
        "Verification Needed": "Confirm current ownership generation; confirm size; not confirmed PE-free",
    },

    # ── SILVER TSUNAMI ADDS — Elevator | Funeral | Calibration | Shredding ──
    {
        "Company Name":       "Mid-American Elevator Company",
        "Website":            "mid-americanelevator.com",
        "State":              "IL",
        "City/HQ":            "East Dundee",
        "Industry":           "Elevator Sales, Service & Modernization",
        "Est. Revenue":       "$8M–$15M",
        "Est. EBITDA":        "$1.2M–$2.5M",
        "Est. Employees":     "50–75",
        "Ownership Type":     "Family (Rob Bailey, LBO 1985)",
        "Year Founded":       "~1972",
        "Multi-Location?":    "Yes (Chicago metro)",
        "Priority Score":     "1",
        "Status":             "Researching",
        "Contact Name":       "Rob Bailey",
        "Title":              "President",
        "Email":              "rbailey@mid-americanelevator.com",
        "Phone":              "(773) 486-6900",
        "LinkedIn URL":       "linkedin.com/company/mid-american-elevator-co/",
        "Last Outreach Date": "",
        "Next Step":          "Call (773) 486-6900. Rob and his father did a leveraged MBO in 1985; Rob is ~74 (Northwestern 1974). Son Robby runs sister company USA Hoist. ZoomInfo confirms r***@mid-americanelevator.com.",
        "Notes / Why Attractive": "Rob Bailey (Northwestern 1974 = ~74 yrs old) did LBO of Mid-American with his father in 1985. Chicago's largest independent elevator co, founded 1974. Son Robby manages USA Hoist (rack & pinion rental sister company). Two-company family elevator platform. ZoomInfo confirms r***@ = rbailey. APi Group and 3Phase actively buying Chicago-area operators.",
        "Verification Needed": "Confirm Rob vs. Cullen as deal decision-maker; confirm no PE approach yet; verify email prefix via call",
    },
    {
        "Company Name":       "Routsong Funeral Home",
        "Website":            "routsong.com",
        "State":              "OH",
        "City/HQ":            "Kettering",
        "Industry":           "Funeral Homes & Death Care",
        "Est. Revenue":       "$3M–$8M",
        "Est. EBITDA":        "$600K–$1.5M",
        "Est. Employees":     "15–30",
        "Ownership Type":     "Family (3rd gen)",
        "Year Founded":       "~1948",
        "Multi-Location?":    "Yes (Kettering + Centerville OH)",
        "Priority Score":     "1",
        "Status":             "Researching",
        "Contact Name":       "Thomas 'Tommy' Routsong III",
        "Title":              "Owner & Funeral Director",
        "Email":              "tommy@routsong.com",
        "Phone":              "(937) 293-4137",
        "LinkedIn URL":       "routsong.com/about-us/",
        "Last Outreach Date": "",
        "Next Step":          "CALL FIRST — (937) 293-4137. Do NOT lead with M&A. Reference 3rd-gen legacy and ask about the business. Park Lawn is very active in Dayton market right now.",
        "Notes / Why Attractive": "3rd-gen Dayton-area funeral home (Kettering + Centerville OH). Tommy Routsong bought from his father Thomas H. Routsong in 2002; born ~1957 (Albion College 1979), now ~68 years old. NFDA member. Park Lawn (TSX: PLC) and SCI both aggressively buying Ohio funeral homes. Strong pre-need contracts. Call before email — relationship approach works here.",
        "Verification Needed": "Confirm Tommy as sole decision-maker; verify email (also try thomas.routsong@routsong.com); check if SCI/Park Lawn already approached",
    },
    {
        "Company Name":       "Geib Funeral Homes & Crematories",
        "Website":            "geibfuneral.com",
        "State":              "OH",
        "City/HQ":            "New Philadelphia",
        "Industry":           "Funeral Homes & Death Care",
        "Est. Revenue":       "$4M–$9M",
        "Est. EBITDA":        "$800K–$1.8M",
        "Est. Employees":     "20–40",
        "Ownership Type":     "Family (5th gen, since 1846)",
        "Year Founded":       "1846",
        "Multi-Location?":    "Yes (Dover + New Philadelphia OH)",
        "Priority Score":     "1",
        "Status":             "Researching",
        "Contact Name":       "Anne Dorris",
        "Title":              "President (5th gen, née Geib)",
        "Email":              "adorris@geibfuneral.com",
        "Phone":              "(330) 364-5538",
        "LinkedIn URL":       "geibfuneral.com/about-us/our-staff",
        "Last Outreach Date": "",
        "Next Step":          "Call New Philadelphia HQ (330) 364-5538; ask for Anne Dorris (née Geib). Father Richard D. Geib II has stepped back — Anne now runs it as 5th gen President. Park Lawn active in Ohio.",
        "Notes / Why Attractive": "One of Ohio's oldest funeral homes (est. 1846). Anne Dorris (née Geib) is President — 5th generation. Father Richard D. Geib II stepped back, now runs consulting firm. 2 locations: Dover and New Philadelphia (Tuscarawas County OH). Pre-need contracts + real estate = strong asset base. Park Lawn (TSX: PLC) most aggressive Ohio buyer right now.",
        "Verification Needed": "Confirm Anne as primary decision-maker; verify email (a.dorris@ also possible); confirm Richard fully retired from operations; check if Park Lawn/SCI approached",
    },
    {
        "Company Name":       "Pacific Shredding",
        "Website":            "pacificshredding.com",
        "State":              "CA",
        "City/HQ":            "Sacramento",
        "Industry":           "Secure Document Destruction / Shredding",
        "Est. Revenue":       "$5M–$12M",
        "Est. EBITDA":        "$800K–$2M",
        "Est. Employees":     "25–50",
        "Ownership Type":     "Division of Pacific Storage Company (family-controlled)",
        "Year Founded":       "1979 (as division of Pacific Storage)",
        "Multi-Location?":    "Yes (Sacramento, Stockton, Modesto CA)",
        "Priority Score":     "2",
        "Status":             "Researching",
        "Contact Name":       "Phil Johnson",
        "Title":              "Chairman, Pacific Storage Company (parent)",
        "Email":              "pjohnson@pacificstorage.com",
        "Phone":              "[verify via pacificstorage.com/board-of-directors]",
        "LinkedIn URL":       "pacificstorage.com/board-of-directors/",
        "Last Outreach Date": "",
        "Next Step":          "This is a carve-out / division sale opportunity. Contact Pacific Storage Company parent to explore whether the shredding division can be separated. Visit pacificstorage.com for contact info. Pacific Shredding ≠ standalone company.",
        "Notes / Why Attractive": "Pacific Shredding is a division of Pacific Storage Company (founded 1856, Sacramento CA). Division since 1979. 3 California markets: Sacramento, Stockton, Modesto. Phil Johnson is Chairman of parent board. Position as strategic divestiture of non-core division — Iron Mountain and ProShred both buying CA route-based shredding books. Tom LaRochelle was NOT the owner — he does not exist.",
        "Verification Needed": "Confirm Phil Johnson is right contact for divestiture; get Pacific Shredding division revenue separately; assess if management would pursue MBO; confirm still operates as independent division",
    },
]

# ── PE-acquired / disqualified ───────────────────────────────────────────────
DISQUALIFIED = [
    # Previously identified
    {
        "Company Name": "Sansone Air Conditioning",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Acquired by Strikepoint Group Holdings (PE), October 2022",
        "Website":      "sansone-ac.com",
        "State":        "FL",
        "Notes":        "Family-owned since 1976 (David Sansone, President). Acquired by PE platform Strikepoint.",
    },
    {
        "Company Name": "Piper Fire Protection",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Acquired by Fortis Fire & Safety (PE), March 2023",
        "Website":      "piperfire.com",
        "State":        "FL",
        "Notes":        "Founded 1986 by Terry Johnson; Chris Johnson was President. Fortis is backed by private equity.",
    },
    {
        "Company Name": "Stan's Heating Air & Plumbing",
        "Category":     "PE-BACKED",
        "Reason":       "Received investment from Treaty Oak Equity (Austin TX)",
        "Website":      "stansac.com",
        "State":        "TX",
        "Notes":        "Chris Strand owner; institutional growth investment. Not a clean independent buyout.",
    },
    # New removals
    {
        "Company Name": "Statewide Fire Corp",
        "Category":     "FOREIGN-ACQUIRED",
        "Reason":       "Acquired by Scutum Group (French security company), operating as Scutum Digital",
        "Website":      "statewidefirecorp.com",
        "State":        "NY",
        "Notes":        "Founded 2002 by Pamela Coppola-Columbia. Now part of Scutum North America. Not a buyout target.",
    },
    {
        "Company Name": "Down to Earth Landscaping",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Acquired by SCG Partners (PE), Florida landscaping roll-up",
        "Website":      "dtelandscape.com",
        "State":        "FL",
        "Notes":        "1,400 employees, 15 locations. SCG Partners PE platform. Too large and PE-backed.",
    },
    {
        "Company Name": "Jack Lehr Heating Cooling & Electric",
        "Category":     "ESOP",
        "Reason":       "100% employee-owned (ESOP) since 2021 — not a buyout candidate",
        "Website":      "jacklehr.com",
        "State":        "PA",
        "Notes":        "Founded 1973 by Jack Lehr; went ESOP in 2021. President Clint Solliday. ESOPs don't sell to PE.",
    },
    {
        "Company Name": "Atlantic Testing Laboratories",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Acquired by Phenna Group (UK PE-backed), announced 2023",
        "Website":      "atlantictesting.com",
        "State":        "NY",
        "Notes":        "Founded 1967 by Spencer Thew; CEO Marijean Remington. 270 employees, 11 offices. Phenna Group = PE-backed testing roll-up.",
    },
    {
        "Company Name": "JM Test Systems",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Acquired by Kanbrick (PE), 2023",
        "Website":      "jmtest.com",
        "State":        "LA",
        "Notes":        "Founded 1982 by Ed Morrison; sons ran it. 10 locations. Kanbrick acquired 2023. New CEO Andrew Treanor.",
    },
    {
        "Company Name": "Nouveau Elevator Industries",
        "Category":     "TOO LARGE",
        "Reason":       "950+ employees, 10,000+ elevators — above lower middle market threshold",
        "Website":      "nouveauelevator.com",
        "State":        "NY",
        "Notes":        "Robert Speranza (Owner), Donald Speranza Sr. (President). One of largest independent elevator cos in US. Not LMM.",
    },
    {
        "Company Name": "Delaware Elevator",
        "Category":     "TOO LARGE",
        "Reason":       "550+ employees, 4th-gen, multi-state — above lower middle market threshold",
        "Website":      "delawareelevator.com",
        "State":        "MD",
        "Notes":        "Charles Meeks (President), Charles E. Meeks Jr. (Key Principal). Founded 1946. Branches MD/DE/VA/Carolinas/FL.",
    },
    {
        "Company Name": "Princeton Air Conditioning",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Acquired by HomeX Services Group, August 13, 2024 (their 17th acquisition)",
        "Website":      "princetonair.com",
        "State":        "NJ",
        "Notes":        "Scott Needham was President. HomeX is PE-backed home services roll-up. Deal closed Aug 2024 — no longer independent.",
    },
    {
        "Company Name": "Fireline Corporation",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Majority recapitalization by Encore Fire Protection, January 16, 2025",
        "Website":      "fireline.com",
        "State":        "MD",
        "Notes":        "Anna Waters Gavin (3rd gen) sold majority to Encore. Chesapeake Corporate Advisors was sell-side advisor. Encore itself sold to Permira (PE) March 2025.",
    },
    {
        "Company Name": "Bruni & Campisi",
        "Category":     "PE-BACKED",
        "Reason":       "Part of Apex Service Partners platform (Alpine Investors PE-backed)",
        "Website":      "bruniandcampisi.com",
        "State":        "NY",
        "Notes":        "Keith Bruni now listed at Apex Service Partners on LinkedIn. Apex operates 107+ brands under original names. Apollo Funds took minority stake in Apex May 2026.",
    },
    {
        "Company Name": "E2B Calibration",
        "Category":     "PE-ACQUIRED",
        "Reason":       "Acquired by Transcat Inc. (NASDAQ: TRNS), September 2022",
        "Website":      "e2bcal.com",
        "State":        "OH",
        "Notes":        "Now operates as Transcat's Cleveland Calibration Lab. ISO 17025 accredited. Was in NE Ohio. Good case study for pitching similar calibration labs to Transcat.",
    },
]


def build_workbook(output_path: str):
    wb = Workbook()

    # ── Sheet 1: Pipeline ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Pipeline"

    for ci, (label, width, wrap, align) in enumerate(PIPELINE_COLS, 1):
        cell = ws.cell(row=1, column=ci, value=label)
        cell.fill      = _fill(NAV)
        cell.font      = _font(bold=True, color=WHITE, size=11)
        cell.alignment = _align(h="center", v="center")
        cell.border    = _border()
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(PIPELINE_ROWS, 2):
        p = row.get("Priority Score", "2")
        if p == "1":
            bg, fg = GREEN_BG, GREEN_FG
        elif p == "2":
            bg, fg = AMBER_BG, AMBER_FG
        else:
            bg, fg = GRAY_BG, GRAY_FG

        for ci, (label, _, wrap, halign) in enumerate(PIPELINE_COLS, 1):
            val  = row.get(label, "")
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(color=fg, size=10)
            cell.border    = _border()
            cell.alignment = _align(h=halign, v="top", wrap=wrap)

            if label == "Email" and val:
                confirmed = "✓" in str(val)
                cell.font = Font(
                    color="1155CC" if not confirmed else "375623",
                    underline="single", size=10, name="Calibri",
                    bold=confirmed
                )
            if label == "LinkedIn URL" and val:
                cell.font = Font(color="0563C1", size=9, name="Calibri")
            if label in ("Notes / Why Attractive", "Next Step"):
                cell.font = _font(italic=True, color=fg, size=9)
            if label == "Priority Score":
                badge = {"1": "375623", "2": "806000", "3": "595959"}.get(p, "595959")
                cell.fill = _fill(badge)
                cell.font = _font(bold=True, color=WHITE, size=10)
                cell.alignment = _align(h="center", v="center")

        ws.row_dimensions[ri].height = 80

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Sheet 2: Disqualified ──────────────────────────────────────────────────
    dq = wb.create_sheet("Disqualified")
    dq_cols = [
        ("Company Name", 26), ("State", 7), ("Category", 18),
        ("Reason", 52), ("Notes", 52),
    ]
    for ci, (label, width) in enumerate(dq_cols, 1):
        cell = dq.cell(row=1, column=ci, value=label)
        cell.fill      = _fill("843C0C")
        cell.font      = _font(bold=True, color=WHITE, size=11)
        cell.alignment = _align(h="center", v="center")
        cell.border    = _border()
        dq.column_dimensions[get_column_letter(ci)].width = width
    dq.row_dimensions[1].height = 28

    CAT_COLORS = {
        "PE-ACQUIRED":     ("FCE4D6", "843C0C"),
        "PE-BACKED":       ("FCE4D6", "843C0C"),
        "FOREIGN-ACQUIRED":("FFF2CC", "7F6000"),
        "ESOP":            ("F2F2F2", "595959"),
        "TOO LARGE":       ("D6E4F7", "1F3864"),
    }
    for ri, row in enumerate(DISQUALIFIED, 2):
        cat = row.get("Category", "PE-ACQUIRED")
        bg, fg = CAT_COLORS.get(cat, ("FCE4D6", "843C0C"))
        for ci, (label, _) in enumerate(dq_cols, 1):
            val  = row.get(label, "")
            cell = dq.cell(row=ri, column=ci, value=val)
            cell.fill      = _fill(bg)
            cell.font      = _font(color=fg, size=10)
            cell.border    = _border()
            cell.alignment = _align(h="left", v="top", wrap=True)
        dq.row_dimensions[ri].height = 56
    dq.freeze_panes = "A2"

    # ── Sheet 3: Quick-Reference Email Sheet ──────────────────────────────────
    er = wb.create_sheet("Email Cheatsheet")
    er.column_dimensions["A"].width = 26
    er.column_dimensions["B"].width = 20
    er.column_dimensions["C"].width = 40
    er.column_dimensions["D"].width = 18
    er.column_dimensions["E"].width = 16

    hdr_vals = ["Company", "Contact", "Email (send here)", "Phone", "Status"]
    for ci, v in enumerate(hdr_vals, 1):
        cell = er.cell(row=1, column=ci, value=v)
        cell.fill = _fill(NAV); cell.font = _font(bold=True, color=WHITE, size=11)
        cell.border = _border(); cell.alignment = _align(h="center")
    er.row_dimensions[1].height = 28

    email_rows = [
        ("── PRIORITY 1 ──", "", "", "", ""),
        ("Aiello Home Services",    "Michael Jezouit",    "m.jezouit@aiellohomeservices.com ✓", "(860) 292-2600", "CONFIRMED"),
        ("Mazza Recycling",         "James Mazza Jr.",    "j.mazza@mazzarecycling.com",         "(732) 992-8055", "Estimated"),
        ("Mowrey Elevator",         "Tim Mowrey Sr.",     "t.mowrey@mowreyelevator.com",        "(850) 526-4111", "Estimated"),
        ("Russell Reid",            "Gary M. Weiner",     "gweiner@russellreid.com",            "(800) 356-4468", "Estimated"),
        ("── VERIFY FIRST (P2) ──", "", "", "", ""),
        ("Burkholder's HVAC",       "Bob Burkholder",     "r.burkholder@burkholders-hvac.com",  "(610) 816-6889", "Estimated"),
        ("Confires Fire",           "W. Wrublevski",      "william.wrublevski@confires.com",    "(908) 822-2700", "Format confirmed"),
        ("T.F. O'Brien",            "Kerry O'Brien",      "k.obrien@tfobrien.com",              "(516) 488-1800", "Estimated"),
        ("Suburban Disposal",       "Chris Roselle",      "c.roselle@suburbandisposalinc.com",  "(973) 227-7020", "Estimated"),
        ("A Royal Flush",           "Tim Butler",         "tbutler@aroyalflush.com",            "(203) 333-1400", "Estimated"),
        ("Texas Outhouse",          "Paul Carl",          "paul.carl@texasouthouse.com",        "(713) 690-9800", "Estimated"),
        ("Revolution Recovery",     "Jonathan Wybar",     "jwybar@revolutionrecovery.com",      "(215) 333-6505", "Estimated"),
        ("Bosworth AC",             "Jerry Bosworth Jr.", "jerry@bosworthac.com",               "(409) 762-0418", "Estimated"),
        ("── SILVER TSUNAMI ADDS ──", "", "", "", ""),
        ("Mid-American Elevator",   "Rob Bailey",          "rbailey@mid-americanelevator.com",   "(773) 486-6900", "ZoomInfo: r***"),
        ("Routsong Funeral Home",   "Tommy Routsong III",  "tommy@routsong.com",                 "(937) 293-4137", "Estimated — call first"),
        ("Geib Funeral Home",       "Anne Dorris",         "adorris@geibfuneral.com",            "(330) 364-5538", "Estimated"),
        ("Pacific Shredding",       "Phil Johnson (parent co.)", "pjohnson@pacificstorage.com",  "[verify]",       "Carve-out: call first"),
    ]

    for ri, (co, name, email, phone, status) in enumerate(email_rows, 2):
        vals = [co, name, email, phone, status]
        is_hdr = co.startswith("──")
        is_confirmed = "✓" in email
        for ci, v in enumerate(vals, 1):
            cell = er.cell(row=ri, column=ci, value=v)
            cell.border = _border()
            if is_hdr:
                cell.fill = _fill(NAV); cell.font = _font(bold=True, color=WHITE, size=10)
                cell.alignment = _align(h="left")
                er.row_dimensions[ri].height = 22
            elif is_confirmed:
                cell.fill = _fill(GREEN_BG)
                if ci == 3:
                    cell.font = Font(color="375623", bold=True, size=10, name="Calibri")
                else:
                    cell.font = _font(color=GREEN_FG, bold=(ci==1), size=10)
                cell.alignment = _align(h="left", v="center")
                er.row_dimensions[ri].height = 20
            else:
                cell.fill = _fill(AMBER_BG)
                cell.font = _font(color=AMBER_FG, size=10)
                cell.alignment = _align(h="left", v="center")
                er.row_dimensions[ri].height = 20
    er.freeze_panes = "A2"

    wb.save(output_path)

    p1 = [r for r in PIPELINE_ROWS if r["Priority Score"] == "1"]
    p2 = [r for r in PIPELINE_ROWS if r["Priority Score"] == "2"]
    p3 = [r for r in PIPELINE_ROWS if r["Priority Score"] == "3"]
    confirmed = [r for r in PIPELINE_ROWS if "✓" in r.get("Email","")]

    print(f"\n✓ Pipeline saved → {output_path}")
    print(f"  Sheet 1 'Pipeline':        {len(PIPELINE_ROWS)} contacts  |  {len(p1)} P1  |  {len(p2)} P2  |  {len(p3)} P3")
    print(f"  Sheet 2 'Disqualified':    {len(DISQUALIFIED)} removed  (PE-backed, ESOP, foreign, too large)")
    print(f"  Sheet 3 'Email Cheatsheet': quick-send reference")
    print(f"\n  CONFIRMED emails ({len(confirmed)}):")
    for r in confirmed:
        print(f"    {r['Company Name']:30} → {r['Email']}")
    print(f"\n  P1 targets (send this week):")
    for r in p1:
        print(f"    {r['Company Name']:30} {r['Contact Name']}")


def main():
    parser = argparse.ArgumentParser(description="Populate M&A pipeline from researched contacts")
    parser.add_argument("--output", default="MA_Pipeline_Populated.xlsx")
    args = parser.parse_args()
    build_workbook(args.output)


if __name__ == "__main__":
    main()
