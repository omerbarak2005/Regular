"""
Project Alpine CIM Generator
Generates Project_Alpine_CIM.xlsx — a full M&A deal package for a 35-unit
Dutch Bros Coffee franchise portfolio in the Pacific Northwest.
Advisor: Jim O'Donnell, MD | Paulson Investment Company
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY        = "1F3864"
WHITE       = "FFFFFF"
GREEN_FILL  = "E2EFDA"
GOLD_FILL   = "FFF2CC"
RED_FILL    = "FCE4D6"
BLUE_FILL   = "DDEEFF"
LIGHT_GREY  = "F2F2F2"
DARK_GREY   = "595959"
GREEN_ROW   = "C6EFCE"   # our deal highlight
RED_ROW     = "FFC7CE"   # Arizona comp warning

# ── Style helpers ───────────────────────────────────────────────────────────────
def hdr_font(bold=True, color=WHITE, size=11):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def body_font(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_bottom():
    thin = Side(style="thin", color="BFBFBF")
    thick = Side(style="medium", color=NAVY)
    return Border(left=thin, right=thin, top=thin, bottom=thick)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def right_align():
    return Alignment(horizontal="right", vertical="center")

def apply_header(cell, text, col_span=None):
    cell.value = text
    cell.font = hdr_font()
    cell.fill = fill(NAVY)
    cell.alignment = center()
    cell.border = thin_border()

def apply_label(cell, text):
    cell.value = text
    cell.font = body_font(bold=True)
    cell.fill = fill(LIGHT_GREY)
    cell.alignment = left()
    cell.border = thin_border()

def apply_value(cell, text):
    cell.value = text
    cell.font = body_font()
    cell.alignment = left()
    cell.border = thin_border()

def apply_section_title(ws, row, col, text, ncols=6):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    cell.fill = fill(NAVY)
    cell.alignment = left()
    cell.border = thin_border()
    if ncols > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + ncols - 1
        )

def freeze_and_filter(ws, freeze_cell="A2"):
    ws.freeze_panes = freeze_cell

# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — Deal Summary
# ═══════════════════════════════════════════════════════════════════════════════
def build_deal_summary(wb):
    ws = wb.create_sheet("Deal Summary")

    # Title banner
    ws.merge_cells("A1:B1")
    title = ws["A1"]
    title.value = "PROJECT ALPINE — CONFIDENTIAL INFORMATION MEMORANDUM"
    title.font = Font(name="Calibri", bold=True, color=WHITE, size=14)
    title.fill = fill(NAVY)
    title.alignment = center()

    ws.merge_cells("A2:B2")
    sub = ws["A2"]
    sub.value = "35-Unit Dutch Bros Coffee Franchise Portfolio | Pacific Northwest | Paulson Investment Company"
    sub.font = Font(name="Calibri", bold=False, color=WHITE, size=11, italic=True)
    sub.fill = fill(NAVY)
    sub.alignment = center()

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    # Column headers
    headers = ["DEAL PARAMETER", "VALUE / DETAIL"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill(NAVY)
        c.alignment = center()
        c.border = thin_border()

    rows = [
        ("Project Name",               "Project Alpine  (CONFIDENTIAL)"),
        ("Asset",                       "35-Unit Dutch Bros Coffee Franchise Portfolio — Pacific Northwest"),
        ("Revenue (TTM)",               "$70M+"),
        ("Adj. EBITDA (TTM)",           "$13.5M"),
        ("EBITDA Margin",               "18.5%"),
        ("Units",                       "35 licensed drive-through locations"),
        ("Expansion Potential",         "5–7 additional sites within 1-mile of existing"),
        ("Corporate Offer",             "4.0x EBITDA (~$54M) — REJECTED"),
        ("Arizona Comp (May 2026)",     "4.5x (29 units) — retirement discount, not market rate"),
        ("Ask",                         "7.0x–8.0x EBITDA  ($94.5M–$108M)"),
        ("NTM EBITDA (w/ new units)",   "~$16.5M  (→ 7x looks like 5.7x forward)"),
        ("ROFR Status",                 "Dutch Bros has ROFR — stated will NOT match 6-8x in writing"),
        ("Strategy",                    "Stalking horse LOI at 7x + 3% break-up fee → forces ROFR decision"),
        ("Advisor",                     "Jim O'Donnell, MD  |  Paulson Investment Company"),
        ("Status",                      "ACTIVE — seeking LOI"),
    ]

    # Alternate fill colors
    alt_fills = [fill("FFFFFF"), fill(LIGHT_GREY)]
    for i, (label, value) in enumerate(rows, start=4):
        row_fill = alt_fills[i % 2]
        lc = ws.cell(row=i, column=1, value=label)
        lc.font = body_font(bold=True, size=10)
        lc.fill = row_fill
        lc.alignment = left()
        lc.border = thin_border()

        vc = ws.cell(row=i, column=2, value=value)
        vc.font = body_font(size=10)
        vc.fill = row_fill
        vc.alignment = left()
        vc.border = thin_border()

        # Highlight key rows
        if label == "Ask":
            lc.font = body_font(bold=True, color="375623", size=10)
            vc.font = body_font(bold=True, color="375623", size=10)
            lc.fill = fill(GREEN_FILL)
            vc.fill = fill(GREEN_FILL)
        elif label == "Corporate Offer":
            lc.fill = fill(RED_FILL)
            vc.fill = fill(RED_FILL)
        elif label == "Arizona Comp (May 2026)":
            lc.fill = fill(RED_FILL)
            vc.fill = fill(RED_FILL)
        elif label == "Status":
            vc.font = Font(name="Calibri", bold=True, color="FF0000", size=10)
        elif label == "Strategy":
            lc.fill = fill(GOLD_FILL)
            vc.fill = fill(GOLD_FILL)

    # Footnote
    fn_row = len(rows) + 4 + 1
    ws.merge_cells(f"A{fn_row}:B{fn_row}")
    fn = ws.cell(row=fn_row, column=1,
                 value="STRICTLY CONFIDENTIAL — This document is intended solely for the named recipient. "
                       "Distribution without written consent of Paulson Investment Company is prohibited.")
    fn.font = Font(name="Calibri", italic=True, size=8, color="7F7F7F")
    fn.alignment = center()

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 68
    ws.freeze_panes = "A4"
    print("  [✓] Sheet 1 — Deal Summary")
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — Financial Model
# ═══════════════════════════════════════════════════════════════════════════════
def build_financial_model(wb):
    ws = wb.create_sheet("Financial Model")

    # ── Section A header ────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = "PROJECT ALPINE — FINANCIAL MODEL"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 26

    apply_section_title(ws, 2, 1, "SECTION A — UNIT ECONOMICS (PER STORE & 35-UNIT PORTFOLIO)", ncols=4)

    sec_a_hdrs = ["Line Item", "% of Revenue", "Per Unit ($)", "35-Unit Total ($)"]
    for col, h in enumerate(sec_a_hdrs, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill("2E4057")
        c.alignment = center()
        c.border = thin_border()

    unit_econ = [
        ("Revenue",                                "100.0%",  2_000_000,   70_000_000),
        ("COGS (coffee, food, supplies)",          " 34.0%",    680_000,   23_800_000),
        ("Gross Profit",                           " 66.0%",  1_320_000,   46_200_000),
        ("Labor",                                  " 30.0%",    600_000,   21_000_000),
        ("Rent / Occupancy",                       "  8.0%",    160_000,    5_600_000),
        ("Corporate Royalties (5% of rev)",        "  5.0%",    100_000,    3_500_000),
        ("Marketing Fund (1% of rev)",             "  1.0%",     20_000,      700_000),
        ("Other OpEx (utilities/insurance/maint)", "  3.5%",     70_000,    2_450_000),
        ("EBITDA",                                 " 18.5%",    370_000,   12_950_000),
        ("Mgmt Fee / G&A",                         "  1.5%",     30_000,    1_050_000),
        ("Adjusted EBITDA",                        " 17.0%",    340_000,   11_900_000),
    ]

    bold_rows = {"Revenue", "Gross Profit", "EBITDA", "Adjusted EBITDA"}
    subtotal_rows = {"Gross Profit", "EBITDA", "Adjusted EBITDA"}

    for i, (item, pct, per_unit, total) in enumerate(unit_econ, start=4):
        is_bold = item in bold_rows
        row_fill_hex = "DDEEFF" if item in subtotal_rows else ("FFFFFF" if i % 2 == 0 else LIGHT_GREY)

        for col, val in enumerate([item, pct, per_unit, total], start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=is_bold, size=10)
            c.fill = fill(row_fill_hex)
            c.border = thin_border()
            if col == 1:
                c.alignment = left()
            elif col == 2:
                c.alignment = center()
            else:
                c.alignment = right_align()
                c.number_format = '#,##0'

    # Adjusted EBITDA note
    note_row = 4 + len(unit_econ)
    ws.merge_cells(f"A{note_row}:D{note_row}")
    n = ws.cell(row=note_row, column=1,
                value="  NOTE: Adjusted EBITDA range $11.9M–$13.5M depending on unit mix. "
                      "Several units exceed $2.2M AUV (above system avg), pulling portfolio total to $13.5M+ range.")
    n.font = Font(name="Calibri", italic=True, size=9, color="595959")
    n.fill = fill(GOLD_FILL)
    n.alignment = left()
    n.border = thin_border()

    # ── Section B ───────────────────────────────────────────────────────────────
    sec_b_start = note_row + 2
    apply_section_title(ws, sec_b_start, 1, "SECTION B — 5-YEAR PROJECTION", ncols=6)

    sec_b_hdrs = ["Year", "Units", "Revenue ($M)", "EBITDA ($M)", "EBITDA Margin", "Cumul. New Units"]
    for col, h in enumerate(sec_b_hdrs, start=1):
        c = ws.cell(row=sec_b_start + 1, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill("2E4057")
        c.alignment = center()
        c.border = thin_border()

    proj = [
        ("Year 1 (Current)", 35, 70.0, 13.5, "19.3%", "0 (baseline)"),
        ("Year 2",           36, 73.4, 14.3, "19.5%", "+1"),
        ("Year 3",           38, 79.0, 15.5, "19.6%", "+3"),
        ("Year 4",           40, 85.1, 16.9, "19.9%", "+5"),
        ("Year 5",           42, 91.6, 18.3, "20.0%", "+7"),
    ]

    for i, row_data in enumerate(proj, start=sec_b_start + 2):
        row_fill_hex = GREEN_FILL if row_data[0] == "Year 1 (Current)" else ("FFFFFF" if i % 2 == 0 else LIGHT_GREY)
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=(row_data[0] == "Year 1 (Current)"), size=10)
            c.fill = fill(row_fill_hex)
            c.border = thin_border()
            c.alignment = center() if col != 1 else left()
            if col in (3, 4):
                c.number_format = '#,##0.0'

    # Assumptions row
    assump_row = sec_b_start + 2 + len(proj)
    ws.merge_cells(f"A{assump_row}:F{assump_row}")
    a = ws.cell(row=assump_row, column=1,
                value="  ASSUMPTIONS: 2% same-store sales growth/year | New units at $1.2M build cost | "
                      "Yr-1 EBITDA/new unit = $200K | Stabilized EBITDA/new unit = $340K")
    a.font = Font(name="Calibri", italic=True, size=9, color="595959")
    a.fill = fill(GOLD_FILL)
    a.alignment = left()
    a.border = thin_border()

    # ── Section C ───────────────────────────────────────────────────────────────
    sec_c_start = assump_row + 2
    apply_section_title(ws, sec_c_start, 1, "SECTION C — VALUATION BRIDGE  (Why 7x Is Fair)", ncols=4)

    sec_c_hdrs = ["Scenario", "EBITDA ($M)", "Multiple", "Enterprise Value ($M)"]
    for col, h in enumerate(sec_c_hdrs, start=1):
        c = ws.cell(row=sec_c_start + 1, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill("2E4057")
        c.alignment = center()
        c.border = thin_border()

    val_bridge = [
        ("Corporate Offer (rejected)",       13.5, "4.0x",  54.0,   RED_FILL,   "Dutch Bros internal — not market"),
        ("Arizona Comp Floor (May 2026)",     13.5, "4.5x",  60.75,  RED_FILL,   "Retirement discount — not market rate"),
        ("Our Ask — Entry",                  13.5, "7.0x",  94.5,   GREEN_FILL, "Target LOI price"),
        ("Forward NTM (Year 2)",             14.3, "7.0x", 100.1,   BLUE_FILL,  "= 6.6x on LTM — compelling entry"),
        ("Public Market Arb (BROS at 26.5x)",13.5,"26.5x", 357.75, GOLD_FILL,  "Buying private at 7x → $292M day-1 gain"),
    ]

    for i, (scen, ebitda, mult, ev, row_fill_hex, note) in enumerate(val_bridge, start=sec_c_start + 2):
        vals = [scen, ebitda, mult, ev]
        for col, val in enumerate(vals, start=1):
            c = ws.cell(row=i, column=col, value=val)
            is_bold = scen == "Our Ask — Entry"
            c.font = body_font(bold=is_bold, size=10)
            c.fill = fill(row_fill_hex)
            c.border = thin_border()
            if col == 1:
                c.alignment = left()
            elif col == 3:
                c.alignment = center()
            else:
                c.alignment = right_align()
                c.number_format = '#,##0.0'
        # Note in col 5
        nc = ws.cell(row=i, column=5, value=note)
        nc.font = Font(name="Calibri", italic=True, size=9, color="595959")
        nc.fill = fill(row_fill_hex)
        nc.border = thin_border()
        nc.alignment = left()

    # Column widths
    widths = {"A": 42, "B": 18, "C": 16, "D": 22, "E": 18, "F": 22}
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A3"
    print("  [✓] Sheet 2 — Financial Model")
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — Comparable Transactions
# ═══════════════════════════════════════════════════════════════════════════════
def build_comps(wb):
    ws = wb.create_sheet("Comparable Transactions")

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "PROJECT ALPINE — COMPARABLE TRANSACTIONS"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 26

    headers = ["Deal", "Acquirer", "Seller", "Year", "EV ($M)", "Units", "EV/EBITDA", "Notes"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill(NAVY)
        c.alignment = center()
        c.border = thin_border()

    comps = [
        ("Carrols Restaurant Group", "Restaurant Brands Intl", "Carrols (NASDAQ)", 2023,
         1000.0, "1,000+ BK units", "~9x",
         "Largest BK franchisee; premium for scale",
         BLUE_FILL),
        ("Del Taco", "Jack in the Box", "Del Taco (NASDAQ)", 2021,
         575.0, "600 units", "~8x",
         "Strategic synergy premium",
         BLUE_FILL),
        ("Denny's Franchise Package", "TriArtisan Capital", "Various Denny's ops", 2019,
         175.0, "~110 units", "~8x",
         "Legacy brand, PE entry",
         BLUE_FILL),
        ("WaBa Grill Franchise Pkg", "Private", "Family operators", 2022,
         48.0, "~70 units", "~6x",
         "Regional QSR mid-market",
         BLUE_FILL),
        ("Arizona Dutch Bros Pkg ⚠", "Dutch Bros Corporate", "Legacy operator", 2026,
         63.0, "29 units", "4.5x",
         "RETIREMENT DISCOUNT — operator needed out; NOT market rate",
         RED_FILL),
        ("★ PROJECT ALPINE (Ask)", "TBD", "Our Client (via Paulson)", 2026,
         94.5, "35 units", "7.0x",
         "Premium: expansion land, top-performing units, PNW exclusivity",
         GREEN_FILL),
    ]

    for i, (deal, acq, seller, yr, ev, units, mult, note, row_fill_hex) in enumerate(comps, start=3):
        is_our_deal = deal.startswith("★")
        is_arizona = "Arizona" in deal
        row_data = [deal, acq, seller, yr, ev, units, mult, note]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=is_our_deal, size=10,
                               color="375623" if is_our_deal else ("9C0006" if is_arizona else "000000"))
            c.fill = fill(row_fill_hex)
            c.border = thin_border()
            if col in (4, 5):
                c.alignment = center()
                if col == 5:
                    c.number_format = '#,##0.0'
            elif col in (6, 7):
                c.alignment = center()
            else:
                c.alignment = left()

    # Legend
    leg_row = 3 + len(comps) + 1
    ws.merge_cells(f"A{leg_row}:H{leg_row}")
    leg = ws.cell(row=leg_row, column=1,
                  value="COLOR KEY:  Blue = comparable deals  |  Green = Project Alpine (our deal)  |  "
                        "Red = Arizona comp (retirement discount, should be excluded from market analysis)")
    leg.font = Font(name="Calibri", italic=True, size=9, color="595959")
    leg.fill = fill(LIGHT_GREY)
    leg.alignment = center()
    leg.border = thin_border()

    footnote_row = leg_row + 1
    ws.merge_cells(f"A{footnote_row}:H{footnote_row}")
    fn = ws.cell(row=footnote_row, column=1,
                 value="Source: Public filings, press releases, industry databases. "
                       "Estimated figures marked with '~'. Arizona deal EV estimated based on 4.5x × $14M unit-level EBITDA proxy.")
    fn.font = Font(name="Calibri", italic=True, size=8, color="7F7F7F")
    fn.fill = fill("FFFFFF")
    fn.alignment = left()

    col_widths = {"A": 30, "B": 28, "C": 26, "D": 8, "E": 12, "F": 14, "G": 12, "H": 52}
    for col_letter, w in col_widths.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A3"
    print("  [✓] Sheet 3 — Comparable Transactions")
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — Buyer Targets
# ═══════════════════════════════════════════════════════════════════════════════
def build_buyer_targets(wb):
    ws = wb.create_sheet("Buyer Targets")

    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "PROJECT ALPINE — BUYER TARGETS & CONTACT DIRECTORY"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 26

    # ── SECTION 1: STRATEGIC OPERATORS ─────────────────────────────────────────
    apply_section_title(ws, 2, 1,
        "PRIORITY 1 — STRATEGIC OPERATORS  (Pay Highest Multiples; Direct Operators)", ncols=9)

    strat_hdrs = ["Buyer", "Type", "Contact", "Title", "Email", "Phone", "Status", "Priority", "Notes"]
    for col, h in enumerate(strat_hdrs, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill("2F5496")
        c.alignment = center()
        c.border = thin_border()

    strategics = [
        ("Flynn Restaurant Group", "Operator", "Lorin Cortina", "EVP & CFO",
         "lcortina@flynnrg.com", "(415) 362-5000", "ACTIVE", "P1",
         "Largest franchise operator in US; Greg Flynn email bounced — use Lorin"),
        ("Sun Holdings", "Operator", "Guillermo Perales", "Founder & CEO",
         "gperales@sunholdings.com", "(214) 452-0015", "ACTIVE", "P1",
         "1,000+ franchise units; can move fast"),
        ("Tucker's Farm", "Regional Op", "Tucker Family", "Owner",
         "[verify via website]", "(541) 555-0100", "ACTIVE", "P1",
         "PNW operator — strong geographic fit"),
    ]

    for i, row_data in enumerate(strategics, start=4):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=(col == 1), size=10)
            c.fill = fill(GREEN_FILL)
            c.border = thin_border()
            c.alignment = left()

    # ── SECTION 2: PE FIRMS ─────────────────────────────────────────────────────
    pe_start = 4 + len(strategics) + 1
    apply_section_title(ws, pe_start, 1,
        "PRIORITY 2 — PE FIRMS  (Restaurant / QSR / Consumer Focus)", ncols=9)

    pe_hdrs = ["Firm", "Contact", "Title", "Email", "Phone", "AUM", "Status", "Priority", "Notes"]
    for col, h in enumerate(pe_hdrs, start=1):
        c = ws.cell(row=pe_start + 1, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill("2F5496")
        c.alignment = center()
        c.border = thin_border()

    pe_firms = [
        ("TriArtisan Capital Partners", "Adam Gray", "Managing Partner",
         "agray@triartisan.com", "(212) 257-1700", "$1.2B", "ACTIVE", "P2",
         "Denny's, TGI Friday's specialist — MOST LIKELY PE buyer"),
        ("Gauge Capital", "Jeff Fronterhouse", "Partner",
         "jfronterhouse@gaugecapital.com", "(214) 665-0900", "$900M", "ACTIVE", "P2",
         "Dallas TX, restaurant/consumer focus"),
        ("L Catterton", "Michael Farello", "Managing Partner",
         "mfarello@lcatterton.com", "(203) 682-8200", "$35B", "ACTIVE", "P2",
         "Consumer PE giant; deep QSR knowledge"),
        ("CapitalSpring", "Brent Esber", "Partner",
         "besber@capitalspring.com", "(212) 230-9700", "$1.5B", "ACTIVE", "P2",
         "Restaurant-only PE — lender + equity"),
        ("Trive Capital", "Conner Searcy", "Associate",
         "csearcy@trivecapital.com", "(214) 269-1040", "$1.5B", "ACTIVE", "P2",
         "Texas PE, food/franchise experience"),
        ("Peak Rock Capital", "M&A Team", "—",
         "info@peakrockcapital.com", "(512) 732-7000", "$3B", "ACTIVE", "P2",
         "Austin TX; QSR portfolio focus"),
        ("Goode Partners", "David Oddi", "Partner",
         "doddi@goodepartners.com", "(646) 722-9455", "$800M", "ACTIVE", "P2",
         "NY restaurant/consumer PE"),
        ("Riverside Company", "M&A Team", "—",
         "info@riversideco.com", "(212) 265-7940", "$13B", "ACTIVE", "P2",
         "Global LMM PE; food & consumer"),
        ("Brentwood Associates", "Corp Dev", "—",
         "info@brentwoodassociates.com", "(310) 477-8560", "$3B", "ACTIVE", "P2",
         "LA-based consumer/retail PE"),
        ("Trivest Partners", "Corp Dev", "—",
         "info@trivest.com", "(305) 858-2200", "$4B", "ACTIVE", "P2",
         "Florida PE; founder-exit specialist"),
        ("Orangewood Partners", "Corp Dev", "—",
         "info@orangewoodpartners.com", "(212) 937-3940", "$1B", "ACTIVE", "P2",
         "NY; food & beverage focus"),
    ]

    for i, row_data in enumerate(pe_firms, start=pe_start + 2):
        is_top = row_data[0] == "TriArtisan Capital Partners"
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=(col == 1 or is_top), size=10,
                               color="375623" if is_top else "000000")
            c.fill = fill(GREEN_FILL if is_top else (LIGHT_GREY if i % 2 == 0 else "FFFFFF"))
            c.border = thin_border()
            c.alignment = left()

    # ── SECTION 3: REJECTED ─────────────────────────────────────────────────────
    rej_start = pe_start + 2 + len(pe_firms) + 1
    apply_section_title(ws, rej_start, 1,
        "REJECTED — DO NOT CONTACT  (Conflict, Strategic Misfit, or Client Direction)", ncols=9)

    rejected = [
        ("Roark Capital",           "—", "—", "—", "—", "—", "REJECTED", "—", "Client direction — do not contact"),
        ("GreyLion",                "—", "—", "—", "—", "—", "REJECTED", "—", "Client direction — do not contact"),
        ("Keystone Capital",        "—", "—", "—", "—", "—", "REJECTED", "—", "Client direction — do not contact"),
        ("Garnett Station Partners","—", "—", "—", "—", "—", "REJECTED", "—", "Client direction — do not contact"),
    ]

    for col, h in enumerate(["Firm", "Contact", "Title", "Email", "Phone", "AUM", "Status", "Priority", "Notes"], start=1):
        c = ws.cell(row=rej_start + 1, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill("2F5496")
        c.alignment = center()
        c.border = thin_border()

    for i, row_data in enumerate(rejected, start=rej_start + 2):
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=(col == 8), size=10,
                               color="9C0006" if col == 7 else "000000")
            c.fill = fill(RED_FILL)
            c.border = thin_border()
            c.alignment = left()

    # ── SECTION 4: FAMILY OFFICES ───────────────────────────────────────────────
    fo_start = rej_start + 2 + len(rejected) + 1
    apply_section_title(ws, fo_start, 1,
        "PRIORITY 1 — FAMILY OFFICES  (Best Fit: Permanent Capital / Cash-Yield Story)", ncols=9)

    fo_hdrs = ["Family Office", "Location", "Contact", "Email", "AUM Est.", "Priority", "Status", "—", "Notes / Pitch Angle"]
    for col, h in enumerate(fo_hdrs, start=1):
        c = ws.cell(row=fo_start + 1, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill("7F6000")
        c.alignment = center()
        c.border = thin_border()

    family_offices = [
        ("Endeavour Capital", "Portland, OR", "Wally Turner (MP)",
         "info@endeavourcapital.com", "$1.5B", "P1", "ACTIVE",
         "★ HOME TURF",
         "PNW-based, consumer focus — geographic alignment ideal"),
        ("Columbia Pacific Advisors", "Seattle, WA", "M&A Team",
         "info@columbiapacific.com", "$800M", "P1", "ACTIVE",
         "",
         "PNW family office; restaurant-familiar"),
        ("Altair Advisers", "Chicago, IL", "Corp Dev",
         "client@altairadvisers.com", "$5B+", "P1", "ACTIVE",
         "",
         "Multi-family office; permanent capital"),
        ("Veritas Financial", "Dallas, TX", "Managing Dir",
         "info@veritasfinancial.com", "$2B", "P1", "ACTIVE",
         "",
         "TX family office; cash-flow focused"),
        ("Lone Star Family Office", "Dallas, TX", "LP Relations",
         "info@lonestarfamilyoffice.com", "$500M+", "P1", "ACTIVE",
         "",
         "TX family; QSR/consumer"),
        ("Pacific Family Capital", "Portland, OR", "Contact",
         "info@pacificfamilycapital.com", "$300M+", "P1", "ACTIVE",
         "",
         "PNW, food/franchise focus"),
        ("Sand Hill Global Advisors", "Palo Alto, CA", "M&A",
         "info@sandhillglobal.com", "$2B+", "P1", "ACTIVE",
         "",
         "Tech-adjacent family office"),
    ]

    for i, row_data in enumerate(family_offices, start=fo_start + 2):
        is_pnw = row_data[1].endswith("OR") or "Portland" in row_data[1]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=(col == 1 or is_pnw), size=10)
            c.fill = fill(GOLD_FILL)
            c.border = thin_border()
            c.alignment = left()

    # Family office pitch box
    pitch_row = fo_start + 2 + len(family_offices) + 1
    apply_section_title(ws, pitch_row, 1, "FAMILY OFFICE PITCH ANGLES", ncols=9)

    pitch_lines = [
        "• $94.5M purchase at 7x → $13.5M EBITDA = 14.3% cash-on-cash yield Year 1",
        "• Levered: 75% LTV debt ($70.9M at 7%) → equity check $23.6M; cash yield on equity ~57% Year 1",
        "• All-cash: $94.5M → $13.5M/year = 14.3% perpetual yield vs. 10-yr Treasury at 4.5%",
        "• Dutch Bros units have 20+ year operating history in Pacific Northwest",
        "• By Year 5, investor collects ~$67.5M cumulative EBITDA on $94.5M purchase — near full payback",
    ]
    for j, line in enumerate(pitch_lines, start=pitch_row + 1):
        ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=9)
        c = ws.cell(row=j, column=1, value=line)
        c.font = body_font(size=10)
        c.fill = fill(GOLD_FILL)
        c.border = thin_border()
        c.alignment = left()

    col_widths_bt = {
        "A": 30, "B": 16, "C": 24, "D": 36,
        "E": 14, "F": 10, "G": 12, "H": 14, "I": 55
    }
    for col_letter, w in col_widths_bt.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A3"
    print("  [✓] Sheet 4 — Buyer Targets")
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — ROFR Strategy
# ═══════════════════════════════════════════════════════════════════════════════
def build_rofr_strategy(wb):
    ws = wb.create_sheet("ROFR Strategy")

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "PROJECT ALPINE — HOW TO GET 7x DESPITE ROFR"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    sub = ws["A2"]
    sub.value = ("The Right of First Refusal (ROFR) held by Dutch Bros Corporate is neutralized by "
                 "a properly structured stalking-horse LOI. Below is the 4-step playbook.")
    sub.font = Font(name="Calibri", italic=True, size=10, color="595959")
    sub.fill = fill(LIGHT_GREY)
    sub.alignment = left()
    sub.border = thin_border()
    ws.row_dimensions[2].height = 32

    steps = [
        {
            "step": "STEP 1",
            "title": "Get ROFR Waiver Statement in Writing",
            "fill": "DDEEFF",
            "lines": [
                "• Corporate Contact: Daniel Warren, EVP Corporate Development, Dutch Bros Inc.",
                "  Email: daniel.warren@dutchbros.com",
                "• Request written confirmation that Dutch Bros will NOT exercise ROFR above 6x EBITDA",
                "• This written statement de-risks the deal entirely for PE buyers and family offices",
                "• Dutch Bros has already publicly stated they will not match 6-8x — get it formalized",
                "• Without this letter, sophisticated buyers will demand a larger break-up fee",
            ],
        },
        {
            "step": "STEP 2",
            "title": "Secure Stalking Horse LOI",
            "fill": GREEN_FILL,
            "lines": [
                "• Primary Target: TriArtisan Capital Partners (Adam Gray, MP — agray@triartisan.com)",
                "• Backup Target: Flynn Restaurant Group (Lorin Cortina, CFO — lcortina@flynnrg.com)",
                "• LOI Terms:",
                "    – Purchase Price: $94.5M (7.0x TTM Adj. EBITDA of $13.5M)",
                "    – Exclusivity Period: 60 days",
                "    – Break-Up Fee: 3% = $2.835M  (paid by SELLER if Dutch Bros exercises ROFR)",
                "• Break-up fee rationale: compensates buyer for DD costs; ensures they engage despite ROFR risk",
                "• Client worst case: Dutch Bros matches → client receives $54M + $2.835M = $56.835M minimum",
                "• Best case: buyer closes at $94.5M → client nets 75% more than corporate offer",
            ],
        },
        {
            "step": "STEP 3",
            "title": "Submit LOI to Dutch Bros Corporate — Force the Decision",
            "fill": GOLD_FILL,
            "lines": [
                "• Per franchise agreement ROFR clause, Dutch Bros has 30 days to match upon receipt of LOI",
                "• Dutch Bros decision tree:",
                "    OPTION A — Match at $94.5M: Pay 7x ($94.5M) → still a compelling deal at their 26.5x public multiple",
                "    OPTION B — Pass: Deal closes to TriArtisan/Flynn → Dutch Bros loses asset",
                "• Dutch Bros at 26.5x public EBITDA multiple (BROS stock): acquiring private EBITDA at 7x",
                "  generates ~$292M in immediate enterprise value arbitrage on Day 1",
                "• If Dutch Bros declines in writing → ROFR is neutralized → clean path to close",
            ],
        },
        {
            "step": "STEP 4",
            "title": "Marketing Timeline — Weeks to Close",
            "fill": LIGHT_GREY,
            "lines": [
                "• Week 1–2:   Execute NDAs with 8–10 qualified buyers from Buyer Target list",
                "• Week 3–5:   Management presentations + data room access",
                "• Week 6–8:   LOI deadline — collect bids; select stalking horse",
                "• Week 9–16:  Exclusivity period + full due diligence",
                "• Week 17–20: Documentation, ROFR notice, closing",
                "  ──────────────────────────────────────────────────────────",
                "• Total estimated timeline: 20 weeks from mandate to close",
                "• Advisor: Jim O'Donnell, MD | Paulson Investment Company",
            ],
        },
    ]

    current_row = 3
    for step_data in steps:
        # Step banner
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        banner = ws.cell(row=current_row, column=1,
                         value=f"  {step_data['step']}:  {step_data['title']}")
        banner.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
        banner.fill = fill(NAVY)
        banner.alignment = left()
        banner.border = thin_border()
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for line in step_data["lines"]:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
            c = ws.cell(row=current_row, column=1, value=line)
            c.font = body_font(size=10)
            c.fill = fill(step_data["fill"])
            c.alignment = Alignment(horizontal="left", vertical="center",
                                    wrap_text=True, indent=1)
            c.border = thin_border()
            ws.row_dimensions[current_row].height = 18
            current_row += 1

        # Spacer
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 30
    ws.freeze_panes = "A3"
    print("  [✓] Sheet 5 — ROFR Strategy")
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — Outreach Tracker
# ═══════════════════════════════════════════════════════════════════════════════
def build_outreach_tracker(wb):
    ws = wb.create_sheet("Outreach Tracker")

    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "PROJECT ALPINE — OUTREACH TRACKER  (Live Deal Management)"
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.fill = fill(NAVY)
    t.alignment = center()
    ws.row_dimensions[1].height = 26

    headers = ["Company", "Contact", "Email", "Outreach Date", "Response", "Status", "Next Step"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = hdr_font(size=10)
        c.fill = fill(NAVY)
        c.alignment = center()
        c.border = thin_border()

    all_targets = [
        # (Company, Contact, Email, Category)
        # Strategics
        ("Flynn Restaurant Group",      "Lorin Cortina",      "lcortina@flynnrg.com",              "Strategic"),
        ("Sun Holdings",                "Guillermo Perales",   "gperales@sunholdings.com",           "Strategic"),
        ("Tucker's Farm",               "Tucker Family",       "[verify via website]",               "Strategic"),
        # PE Firms
        ("TriArtisan Capital Partners", "Adam Gray",           "agray@triartisan.com",              "PE"),
        ("Gauge Capital",               "Jeff Fronterhouse",   "jfronterhouse@gaugecapital.com",    "PE"),
        ("L Catterton",                 "Michael Farello",     "mfarello@lcatterton.com",           "PE"),
        ("CapitalSpring",               "Brent Esber",         "besber@capitalspring.com",          "PE"),
        ("Trive Capital",               "Conner Searcy",       "csearcy@trivecapital.com",          "PE"),
        ("Peak Rock Capital",           "M&A Team",            "info@peakrockcapital.com",          "PE"),
        ("Goode Partners",              "David Oddi",          "doddi@goodepartners.com",           "PE"),
        ("Riverside Company",           "M&A Team",            "info@riversideco.com",              "PE"),
        ("Brentwood Associates",        "Corp Dev",            "info@brentwoodassociates.com",      "PE"),
        ("Trivest Partners",            "Corp Dev",            "info@trivest.com",                  "PE"),
        ("Orangewood Partners",         "Corp Dev",            "info@orangewoodpartners.com",       "PE"),
        # Family Offices
        ("Endeavour Capital",           "Wally Turner",        "info@endeavourcapital.com",         "FO"),
        ("Columbia Pacific Advisors",   "M&A Team",            "info@columbiapacific.com",          "FO"),
        ("Altair Advisers",             "Corp Dev",            "client@altairadvisers.com",         "FO"),
        ("Veritas Financial",           "Managing Dir",        "info@veritasfinancial.com",         "FO"),
        ("Lone Star Family Office",     "LP Relations",        "info@lonestarfamilyoffice.com",     "FO"),
        ("Pacific Family Capital",      "Contact",             "info@pacificfamilycapital.com",     "FO"),
        ("Sand Hill Global Advisors",   "M&A",                 "info@sandhillglobal.com",           "FO"),
    ]

    category_fills = {
        "Strategic": GREEN_FILL,
        "PE":        LIGHT_GREY,
        "FO":        GOLD_FILL,
    }

    for i, (company, contact, email, cat) in enumerate(all_targets, start=3):
        row_fill = category_fills[cat]
        row_data = [
            company, contact, email,
            "",               # Outreach Date — blank for advisor to fill
            "",               # Response
            "Not Yet Contacted",
            "Send NDA + Teaser",
        ]
        for col, val in enumerate(row_data, start=1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = body_font(bold=(col == 1), size=10,
                               color="9C0006" if val == "Not Yet Contacted" else "000000")
            c.fill = fill(row_fill)
            c.border = thin_border()
            c.alignment = left()

    # Legend
    leg_row = 3 + len(all_targets) + 1
    ws.merge_cells(f"A{leg_row}:G{leg_row}")
    leg = ws.cell(row=leg_row, column=1,
                  value=("COLOR KEY:  Green = Strategic Operators (P1)  |  "
                         "Gold = Family Offices (P1)  |  Grey = PE Firms (P2)"))
    leg.font = Font(name="Calibri", italic=True, size=9, color="595959")
    leg.fill = fill(LIGHT_GREY)
    leg.alignment = center()
    leg.border = thin_border()

    col_widths_ot = {
        "A": 30, "B": 24, "C": 36,
        "D": 16, "E": 22, "F": 20, "G": 28
    }
    for col_letter, w in col_widths_ot.items():
        ws.column_dimensions[col_letter].width = w

    ws.freeze_panes = "A3"
    print("  [✓] Sheet 6 — Outreach Tracker")
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    output_path = "/home/user/Regular/Project_Alpine_CIM.xlsx"
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    print("\nBuilding Project_Alpine_CIM.xlsx...")
    print("=" * 55)

    ws1 = build_deal_summary(wb)
    ws2 = build_financial_model(wb)
    ws3 = build_comps(wb)
    ws4 = build_buyer_targets(wb)
    ws5 = build_rofr_strategy(wb)
    ws6 = build_outreach_tracker(wb)

    wb.save(output_path)

    print("=" * 55)
    print(f"\n[COMPLETE] File saved: {output_path}")
    print("\nSHEET SUMMARY:")
    for ws in wb.worksheets:
        print(f"  • {ws.title:<30}  {ws.max_row} rows x {ws.max_column} cols")

    print("\nDATA NOTES / ESTIMATES FLAGGED:")
    print("  ~ Tucker's Farm phone (541) 555-0100 — placeholder; verify via website")
    print("  ~ Arizona Dutch Bros EV ~$63M — ESTIMATED (4.5x × ~$14M EBITDA proxy)")
    print("  ~ BROS public EBITDA multiple of 26.5x — based on trailing data; verify current")
    print("  ~ Peak Rock / Riverside / Brentwood general info emails — verify direct contacts")
    print("  ~ Outreach Dates column is BLANK — advisor to fill in actual send dates")
    print("  ~ Tucker's Farm email listed as [verify via website] — no public email found")
    print("  ~ 10-yr Treasury rate of 4.5% used in family office pitch — verify current rate")
    print("\nAll 6 sheets created successfully.\n")


if __name__ == "__main__":
    main()
