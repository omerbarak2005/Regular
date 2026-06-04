"""
PE Buyers Master Database — Paulson Investment Company
Generates PE_Master_Database.xlsx with 6 sheets covering all active deal sectors.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.filters import AutoFilter

# ── Color palette ─────────────────────────────────────────────────────────────
NAVY       = "1F3864"   # header background
WHITE      = "FFFFFF"
GREEN_BG   = "E2EFDA"   # ✓ Confirmed
AMBER_BG   = "FFF2CC"   # ⚡ High confidence
GRAY_BG    = "F2F2F2"   # ○ Domain only
DARK_GREEN = "375623"   # confirmed font
DARK_AMBER = "7D6608"   # high-conf font
DARK_GRAY  = "595959"   # domain font
LIGHT_BLUE = "DCE6F1"   # alternating sub-rows (unused but available)
SECTION_BG = "D6E4F0"   # section divider
GOLD       = "C9A84C"   # accent on firm name

def navy_fill():
    return PatternFill("solid", fgColor=NAVY)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font():
    return Font(name="Calibri", bold=True, color=WHITE, size=10)

def body_font(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def confidence_style(conf_text):
    """Return (fill, font) based on confidence label."""
    if "Confirmed" in conf_text:
        return make_fill(GREEN_BG), body_font(bold=True, color=DARK_GREEN)
    elif "High" in conf_text:
        return make_fill(AMBER_BG), body_font(bold=False, color=DARK_AMBER)
    else:
        return make_fill(GRAY_BG), body_font(bold=False, color=DARK_GRAY)


def apply_header_row(ws, headers, row=1, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.fill = navy_fill()
        c.font = header_font()
        c.alignment = center()
        c.border = thin_border()


def style_data_row(ws, row_num, num_cols, conf_col_idx=None, conf_value=None):
    """Apply fill based on confidence to the entire row, borders to all cells."""
    if conf_value and conf_col_idx:
        row_fill, _ = confidence_style(conf_value)
    else:
        row_fill = make_fill(WHITE)

    for col in range(1, num_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.border = thin_border()
        c.alignment = left()
        if c.value is None:
            c.value = ""
        # Apply confidence-based fill only if cell doesn't override
        if conf_value and conf_col_idx:
            fill, fnt = confidence_style(conf_value)
            c.fill = fill
            # Bold the buyer name (col 1) always
            if col == 1:
                c.font = body_font(bold=True, color="1F3864", size=10)
            elif col == conf_col_idx:
                c.font = fnt
            else:
                c.font = body_font(size=10)
        else:
            c.fill = row_fill
            if col == 1:
                c.font = body_font(bold=True, color="1F3864", size=10)
            else:
                c.font = body_font(size=10)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def freeze_and_filter(ws, freeze_cell="A2", filter_range=None):
    ws.freeze_panes = freeze_cell
    if filter_range:
        ws.auto_filter.ref = filter_range


def add_sheet_title(ws, title, subtitle, num_cols):
    """Merge top rows for a title banner."""
    ws.insert_rows(1, 2)
    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    tc = ws.cell(row=1, column=1, value=title)
    tc.fill = make_fill(NAVY)
    tc.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Subtitle row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    sc = ws.cell(row=2, column=1, value=subtitle)
    sc.fill = make_fill("2E5F9E")
    sc.font = Font(name="Calibri", bold=False, color=WHITE, size=9, italic=True)
    sc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — QSR & RESTAURANT PE
# ══════════════════════════════════════════════════════════════════════════════

QSR_HEADERS = [
    "Buyer", "Type", "Key Contact", "Title",
    "Email", "Confidence", "Phone", "AUM",
    "Deal Range", "Notes"
]

QSR_DATA = [
    ["TriArtisan Capital Partners", "PE — Restaurant", "Adam Gray", "Managing Partner",
     "agray@triartisan.com", "⚡ High confidence", "(212) 257-1700", "$1.2B",
     "$20M–$500M", "Denny's, TGI Fridays expert. Most likely PE buyer for Dutch Bros."],

    ["Gauge Capital", "PE", "Jeff Fronterhouse", "Partner",
     "jfronterhouse@gaugecapital.com", "⚡ High confidence", "(214) 665-0900", "$900M est.",
     "$50M–$300M", "Dallas TX, restaurant/consumer focus"],

    ["L Catterton", "PE — Consumer", "Michael Farello", "Managing Partner",
     "mfarello@lcatterton.com", "⚡ High confidence", "(203) 682-8200", "$35B",
     "$100M–$1B", "Largest consumer PE globally. Knows QSR unit economics."],

    ["CapitalSpring", "PE — Restaurant Finance", "Brent Esber", "Partner",
     "besber@capitalspring.com", "⚡ High confidence", "(212) 230-9700", "$1.5B",
     "$20M–$200M", "Restaurant-only credit + equity. 300+ deals."],

    ["Trive Capital", "PE", "Conner Searcy", "Partner",
     "csearcy@trivecapital.com", "⚡ High confidence", "(214) 269-1040", "$1.5B",
     "$50M–$300M", "Dallas TX, food/franchise track record"],

    ["Peak Rock Capital", "PE", "M&A Team", "—",
     "info@peakrockcapital.com", "○ Domain only / call first", "(512) 732-7000", "$3B",
     "$50M–$500M", "Austin TX, QSR and consumer portfolio"],

    ["Goode Partners", "PE", "David Oddi", "Co-Founder",
     "doddi@goodepartners.com", "✓ Confirmed", "(646) 722-9455", "$800M est.",
     "$50M–$300M", "NY consumer/restaurant PE. ZoomInfo + phone confirmed."],

    ["Riverside Company", "PE", "M&A Team", "—",
     "info@riversideco.com", "○ Domain only / call first", "(212) 265-7940", "$13B",
     "$10M–$250M", "Global LMM PE, food and consumer portfolio"],

    ["Brentwood Associates", "PE", "M&A Team", "—",
     "info@brentwoodassociates.com", "○ Domain only / call first", "(310) 477-8560", "$3B",
     "$50M–$500M", "LA-based consumer/retail PE"],

    ["Trivest Partners", "PE", "M&A Team", "—",
     "info@trivest.com", "○ Domain only / call first", "(305) 858-2200", "$4B",
     "$20M–$200M", "Founder-exit specialist, Florida HQ"],

    ["Orangewood Partners", "PE", "M&A Team", "—",
     "info@orangewoodpartners.com", "○ Domain only / call first", "(212) 937-3940", "$1B",
     "$50M–$250M", "NY, food & beverage focus"],

    ["Center Bridge Partners", "PE — Large Cap", "Corp Dev", "—",
     "info@centerbridge.com", "○ Domain only / call first", "—", "$35B",
     "$100M+", "Large cap PE, select restaurant/consumer"],

    ["Blackstone Tactical Opps", "PE — Opportunistic", "Corp Dev", "—",
     "tacops@blackstone.com", "○ Domain only / call first", "—", "$300B+",
     "$100M+", "Situation-specific; ROFR complexity could be their play"],

    ["Sun Holdings", "Strategic Operator", "Guillermo Perales", "CEO",
     "gperales@sunholdings.com", "⚡ High confidence", "(214) 452-0015", "—",
     "$50M–$500M", "1,000+ franchise units, fastest-growing US franchisee"],

    ["Flynn Restaurant Group", "Strategic Operator", "Lorin Cortina", "EVP & CFO",
     "lcortina@flynnrg.com", "⚡ High confidence", "(415) 362-5000", "—",
     "$100M+", "Largest US franchise operator. Greg Flynn email bounced — Lorin runs M&A."],

    ["Eyas Capital", "PE — Franchise", "Corp Dev", "—",
     "info@eyascapital.com", "○ Domain only / call first", "—", "est. $500M",
     "$20M–$150M", "Franchise specialist PE"],
]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — HOME SERVICES PE
# ══════════════════════════════════════════════════════════════════════════════

HOME_HEADERS = [
    "Buyer", "Type", "Contact", "Email",
    "Confidence", "Phone", "Deal Range", "Notes"
]

HOME_DATA = [
    ["Apex Service Partners", "PE (Alpine Investors)", "Corp Dev",
     "info@apexservicepartners.com", "○ Domain only / call first", "—",
     "$10M–$100M", "107+ home service brands. Acquired Bruni & Campisi (our disqualified)."],

    ["HomeX Services Group", "PE", "Corp Dev",
     "info@homexservices.com", "○ Domain only / call first", "—",
     "$5M–$50M", "NE expansion. Acquired Princeton Air Aug 2024."],

    ["Wrench Group", "PE (L Catterton)", "Corp Dev",
     "info@wrenchgroup.com", "○ Domain only / call first", "—",
     "$20M–$200M", "HVAC/plumbing rollup. L Catterton backed."],

    ["Neighborly", "PE (KKR/Harvest)", "Corp Dev",
     "info@neighborly.com", "○ Domain only / call first", "—",
     "$10M–$100M", "Mr. Rooter, Aire Serv, Molly Maid. Largest home services franchisor."],

    ["Authority Brands", "PE (Warburg Pincus)", "Corp Dev",
     "info@authoritybrands.com", "○ Domain only / call first", "—",
     "$10M–$100M", "Pest, HVAC, plumbing brands"],

    ["Ember Infrastructure", "PE", "Corp Dev",
     "info@emberinfrastructure.com", "○ Domain only / call first", "—",
     "$20M–$200M", "Home services infrastructure play"],

    ["EMCOR Group", "Public Strategic", "Corp Dev",
     "corpdev@emcor.com", "○ Domain only / call first", "(203) 849-7800",
     "$50M+", "Mechanical + electrical services acquisitions"],

    ["Comfort Systems USA", "Public Strategic", "Corp Dev",
     "info@comfortsystemsusa.com", "○ Domain only / call first", "(713) 830-9600",
     "$30M+", "NYSE: FIX, HVAC roll-up"],
]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — ENVIRONMENTAL & WASTE PE
# ══════════════════════════════════════════════════════════════════════════════

ENV_HEADERS = [
    "Buyer", "Type", "Contact", "Email",
    "Confidence", "Phone", "Deal Range", "Notes"
]

ENV_DATA = [
    ["Pye-Barker Fire & Safety", "PE Rollup", "M&A Team",
     "acquisitions@pyebarkerfs.com", "✓ Confirmed", "—",
     "$5M–$60M", "57 acquisitions in 2024. Most active fire safety buyer on earth."],

    ["Fortis Fire & Safety", "PE Rollup", "M&A Team",
     "info@fortisfire.com", "○ Domain only / call first", "—",
     "$5M–$40M", "Acquired Piper Fire (FL). Expanding."],

    ["Koorsen Fire & Security", "PE Rollup", "Acquisitions",
     "info@koorsen.com", "○ Domain only / call first", "—",
     "$3M–$30M", "Midwest-based, expanding nationally"],

    ["APi Group", "Public Strategic", "Corp Dev",
     "investor.relations@apigroup.com", "✓ Confirmed", "—",
     "$50M+", "NYSE: APG. Owns Elevated Facility Services (elevators)."],

    ["Convergint Technologies", "PE Rollup", "M&A",
     "info@convergint.com", "○ Domain only / call first", "—",
     "$10M–$100M", "Life safety + security integration"],

    ["GFL Environmental", "Public Strategic", "Corp Dev",
     "acquisitions@gflenv.com", "⚡ High confidence", "(905) 326-0101",
     "$20M+", "NYSE: GFL. Active waste acquirer."],

    ["Clean Harbors", "Public Strategic", "Corp Dev",
     "corpdev@cleanharbors.com", "⚡ High confidence", "(781) 792-5000",
     "$30M+", "NYSE: CLH. Hazmat + liquid waste."],

    ["Casella Waste Systems", "Public Strategic", "Corp Dev",
     "acquisitions@casella.com", "⚡ High confidence", "(802) 772-2239",
     "$20M+", "NASDAQ: CWST. NE focus."],

    ["United Site Services", "PE (GI Partners)", "Corp Dev",
     "info@unitedsiteservices.com", "○ Domain only / call first", "—",
     "$5M–$50M", "Portable sanitation rollup. Most aggressive buyer."],

    ["Clean Earth (Harsco)", "PE Rollup", "Corp Dev",
     "info@cleanearthcapital.com", "○ Domain only / call first", "—",
     "$10M–$80M", "Specialty waste disposal."],

    ["Synagro Technologies", "PE Rollup", "Acquisitions",
     "info@synagro.com", "○ Domain only / call first", "—",
     "$5M–$50M", "Biosolids + liquid waste"],
]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — SPECIALTY SERVICES PE
# ══════════════════════════════════════════════════════════════════════════════

SPEC_HEADERS = [
    "Buyer", "Sector", "Contact", "Email",
    "Confidence", "Phone", "Deal Range", "Notes"
]

SPEC_DATA = [
    ["APi Group / Elevated FS", "Elevator", "Corp Dev",
     "investor.relations@apigroup.com", "✓ Confirmed", "—",
     "$20M+", "Paid $570M for Elevated Facility Services 2024. Now buying add-ons."],

    ["3Phase Elevator", "Elevator", "Acquisitions",
     "info@3phaseelevator.com", "○ Domain only / call first", "—",
     "$5M–$30M", "Active independent elevator rollup"],

    ["Schindler Group", "Elevator", "M&A USA",
     "us.info@schindler.com", "○ Domain only / call first", "—",
     "$20M+", "Swiss strategic, pays for market share"],

    ["Transcat Inc.", "Calibration", "Lee Rudow (CEO)",
     "lrudow@transcat.com", "✓ Confirmed", "(585) 352-7777",
     "$5M–$50M", "NASDAQ: TRNS. 26 acquisitions. Most active cal lab buyer."],

    ["Trescal", "Calibration", "Tom Mathews (CEO US)",
     "info@trescal.com", "✓ Confirmed", "—",
     "$5M–$40M", "European PE-backed. Aggressively expanding US."],

    ["Bureau Veritas", "Calibration", "Corp Dev",
     "corpdev@bureauveritas.com", "⚡ High confidence", "—",
     "$20M+", "French-listed. Active US cal lab buyer."],

    ["Intertek", "Calibration", "Corp Dev",
     "corpdev@intertek.com", "⚡ High confidence", "—",
     "$20M+", "London-listed. Buying US testing labs."],

    ["Park Lawn Corporation", "Funeral Homes", "J. Bradley Green",
     "info@parklawn.com", "⚡ High confidence", "(647) 352-3223",
     "$3M–$30M", "TSX: PLC. Most aggressive US funeral home buyer right now."],

    ["Service Corporation Intl", "Funeral Homes", "Corp Dev",
     "corpdev@sci-us.com", "⚡ High confidence", "(713) 522-5141",
     "$5M+", "NYSE: SCI. Largest death care company."],

    ["Carriage Services", "Funeral Homes", "Corp Dev",
     "corpdev@carriageservices.com", "⚡ High confidence", "(832) 429-8002",
     "$3M–$20M", "NYSE: CSV. Active mid-size acquirer."],

    ["Foundation Partners Group", "Funeral Homes", "Acquisitions",
     "info@foundationpartners.com", "○ Domain only / call first", "—",
     "$3M–$20M", "Stonehenge Partners-backed"],

    ["NorthStar Memorial Group", "Funeral Homes", "Corp Dev",
     "info@northstarmemorial.com", "○ Domain only / call first", "—",
     "$5M–$30M", "PE-backed regional rollup"],

    ["Shred-it (Stericycle/WM)", "Shredding", "Corp Dev",
     "corpdev@shred-it.com", "⚡ High confidence", "—",
     "$10M+", "Largest, wants 50+ truck fleets only"],

    ["Iron Mountain", "Shredding", "Corp Dev",
     "corpdev@ironmountain.com", "⚡ High confidence", "(617) 535-4766",
     "$5M+", "NYSE: IRM. Buys smaller regional operators."],

    ["BroadStreet Partners", "Insurance", "Ryan Dobratz",
     "rdobratz@broadstreetpartners.com", "✓ Confirmed", "—",
     "$3M–$50M", "Most active LMM insurance buyer."],

    ["Acrisure", "Insurance", "M&A Team",
     "info@acrisure.com", "○ Domain only / call first", "—",
     "$5M+", "$23B AUM clients. Very aggressive."],

    ["AssuredPartners", "Insurance", "Corp Dev",
     "info@assuredpartners.com", "○ Domain only / call first", "—",
     "$5M–$100M", "GTCR-backed"],

    ["Hub International", "Insurance", "Corp Dev",
     "info@hubinternational.com", "○ Domain only / call first", "—",
     "$10M+", "Hellman & Friedman backed. 10-12x payer."],
]

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — FAMILY OFFICES (Dutch Bros / Project Alpine)
# ══════════════════════════════════════════════════════════════════════════════

FO_HEADERS = [
    "Family Office", "HQ", "Focus", "Contact",
    "Email", "Confidence", "AUM Est.", "Pitch Notes"
]

FO_DATA = [
    ["Endeavour Capital", "Portland OR", "Consumer/PE", "Wally Turner (MP)",
     "info@endeavourcapital.com", "○ Domain only / call first", "$1.5B est.",
     "PNW-based. Geographic fit. Consumer/brand investments. HIGHEST PRIORITY for Dutch Bros."],

    ["Columbia Pacific Advisors", "Seattle WA", "Diversified", "M&A Team",
     "info@columbiapacific.com", "○ Domain only / call first", "$1B est.",
     "Seattle family office. Pacific NW consumer familiarity."],

    ["Altair Advisers", "Chicago IL", "Multi-family office", "Corp Dev",
     "client@altairadvisers.com", "○ Domain only / call first", "$6B+",
     "Large MFO. Has clients who want cash-yield alternatives to bonds."],

    ["Threshold Group", "Seattle WA", "Multi-family office", "Client Relations",
     "info@thresholdgroup.com", "○ Domain only / call first", "$5B+",
     "Seattle MFO. Manages PNW UHNW families. QSR fits their consumer allocation."],

    ["Verdant Enterra Capital", "San Francisco CA", "Consumer/Food", "M&A Team",
     "info@verdantenterra.com", "○ Domain only / call first", "est. $500M",
     "Food-focused family office."],

    ["Cresset Capital", "Chicago IL", "Multi-family office", "Corp Dev",
     "info@cressetcapital.com", "○ Domain only / call first", "$50B+",
     "Top-tier MFO. Direct investment program."],

    ["Baird Family Office Services", "Milwaukee WI", "Multi-family office", "Alternatives",
     "info@rwbaird.com", "○ Domain only / call first", "$250B+",
     "RW Baird direct investments. Conservative but cash-yield focused."],

    ["Laird Norton Wealth Mgmt", "Seattle WA", "Multi-family office", "Alts team",
     "info@lairdnorton.com", "○ Domain only / call first", "$10B+",
     "Seattle PNW MFO. Timber + real assets + private equity allocations."],
]

PITCH_TALKING_POINTS = (
    "PITCH DECK TALKING POINTS — Project Alpine / Dutch Bros 35-Unit:\n\n"
    "At $94.5M entry: $13.5M/yr cash flow = 14.3% yield. 10-yr Treasury: 4.5%. "
    "You're getting 3.2x the bond yield on a business that's been running 20+ years "
    "in the Pacific Northwest and is growing 5–7 new locations. Dutch Bros corporate "
    "has STATED in writing they will not match 6–8x. You hold, you collect $67M over "
    "5 years, then sell at a higher multiple as the unit count grows to 42."
)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — HIGH-GROWTH SECTOR BUYERS
# ══════════════════════════════════════════════════════════════════════════════

HG_HEADERS = [
    "Buyer", "Sector", "Contact", "Email",
    "Confidence", "Phone", "Deal Range", "Notes"
]

HG_DATA = [
    ["Rollins Inc.", "Pest Control", "Acquisitions",
     "acquisitions@rollins.com", "✓ Confirmed", "(404) 888-2000",
     "$3M–$100M", "NYSE: ROL. 40-50 acquisitions/year. Most active pest buyer."],

    ["Rentokil North America", "Pest Control", "Corp Dev",
     "na.corpdev@rentokil.com", "⚡ High confidence", "—",
     "$5M+", "Post-Terminix merger. Very active."],

    ["Arrow Exterminators", "Pest Control", "Acquisitions",
     "info@arrowexterminators.com", "○ Domain only / call first", "—",
     "$3M–$30M", "Private. Family-owned but very active acquirer."],

    ["Anticimex", "Pest Control", "M&A",
     "info@anticimex.com", "○ Domain only / call first", "—",
     "$5M–$50M", "Swedish PE-backed. Expanding US aggressively."],

    ["Blue Leaf Equity", "Pool Service", "M&A",
     "info@blueleafequity.com", "○ Domain only / call first", "—",
     "$3M–$30M", "Pool service rollup. FL/TX/AZ focus."],

    ["Patriot Pool & Spa", "Pool Service", "Acquisitions",
     "info@patriotpool.com", "○ Domain only / call first", "—",
     "$2M–$20M", "Growing PE-backed platform"],

    ["Ntiva", "MSP/IT", "Acquisitions",
     "acquisitions@ntiva.com", "⚡ High confidence", "—",
     "$3M–$30M", "Active MSP buyer. Mid-Atlantic + national."],

    ["Logically", "MSP/IT", "M&A",
     "info@logically.ai", "○ Domain only / call first", "—",
     "$3M–$20M", "New England MSP rollup"],

    ["Ontinue", "MSP/IT", "Corp Dev",
     "info@ontinue.com", "○ Domain only / call first", "—",
     "$5M–$50M", "Cybersecurity-focused MSP buyer"],

    ["Abacus Group", "MSP/IT", "Acquisitions",
     "info@abacusgroup.com", "○ Domain only / call first", "—",
     "$5M–$40M", "Financial services MSP specialist"],

    ["Hopebridge", "ABA Therapy", "Corp Dev",
     "corpdev@hopebridge.com", "⚡ High confidence", "—",
     "$3M–$30M", "Largest ABA provider. Acquisitive."],

    ["Invo Healthcare", "ABA Therapy", "M&A",
     "info@invohealthcare.com", "○ Domain only / call first", "—",
     "$2M–$20M", "Behavioral health services"],

    ["Cortica", "ABA/Behavioral", "M&A",
     "info@corticacare.com", "○ Domain only / call first", "—",
     "$5M–$40M", "PE-backed behavioral health"],

    ["BrightView Holdings", "Landscaping", "Corp Dev",
     "info@brightview.com", "○ Domain only / call first", "(484) 205-5300",
     "$10M+", "NYSE: BV. Commercial landscaping rollup."],

    ["Caliber Collision", "Auto Body", "Corp Dev",
     "info@calibercollision.com", "○ Domain only / call first", "—",
     "$5M–$50M", "Largest collision repair rollup. PE-backed (Hellman & Friedman)."],
]


# ══════════════════════════════════════════════════════════════════════════════
# MAIN BUILD FUNCTION
# ══════════════════════════════════════════════════════════════════════════════

def build_standard_sheet(wb, sheet_name, title_text, subtitle_text,
                          headers, data, conf_col_idx, col_widths):
    """
    Generic builder for sheets that follow the standard layout:
    title banner / subtitle / header row / data rows.
    conf_col_idx: 1-based index of the "Confidence" column.
    """
    ws = wb.create_sheet(sheet_name)
    num_cols = len(headers)

    # Write headers at row 1
    apply_header_row(ws, headers, row=1)

    # Write data rows starting at row 2
    for r_offset, row_data in enumerate(data):
        r = 2 + r_offset
        conf_value = row_data[conf_col_idx - 1] if conf_col_idx <= len(row_data) else ""
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
        style_data_row(ws, r, num_cols, conf_col_idx, conf_value)

    ws.row_dimensions[1].height = 18

    # Add title banner (inserts 2 rows above current row 1)
    add_sheet_title(ws, title_text, subtitle_text, num_cols)
    # Headers are now at row 3 after insert
    freeze_and_filter(ws, freeze_cell="A4",
                      filter_range=f"A3:{get_column_letter(num_cols)}3")

    set_col_widths(ws, col_widths)

    # Row heights for data
    total_rows = 3 + len(data)
    for r in range(4, total_rows + 1):
        ws.row_dimensions[r].height = 40

    return ws, len(data)


def build_family_office_sheet(wb, sheet_name):
    """Special sheet for Family Offices with pitch talking points appended."""
    ws = wb.create_sheet(sheet_name)
    headers = FO_HEADERS
    data = FO_DATA
    num_cols = len(headers)
    conf_col_idx = 6

    apply_header_row(ws, headers, row=1)

    for r_offset, row_data in enumerate(data):
        r = 2 + r_offset
        conf_value = row_data[conf_col_idx - 1]
        for c_idx, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c_idx, value=val)
        style_data_row(ws, r, num_cols, conf_col_idx, conf_value)

    ws.row_dimensions[1].height = 18

    # Blank separator row
    sep_row = 2 + len(data) + 1
    for col in range(1, num_cols + 1):
        ws.cell(row=sep_row, column=col, value="").fill = make_fill(NAVY)

    # Pitch talking points banner
    pitch_row = sep_row + 1
    ws.merge_cells(start_row=pitch_row, start_column=1,
                   end_row=pitch_row, end_column=num_cols)
    pt_cell = ws.cell(row=pitch_row, column=1, value="PROJECT ALPINE — PITCH TALKING POINTS")
    pt_cell.fill = make_fill("375623")
    pt_cell.font = Font(name="Calibri", bold=True, color=WHITE, size=11)
    pt_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[pitch_row].height = 18

    # Pitch text
    body_row = pitch_row + 1
    ws.merge_cells(start_row=body_row, start_column=1,
                   end_row=body_row, end_column=num_cols)
    pb_cell = ws.cell(row=body_row, column=1, value=PITCH_TALKING_POINTS)
    pb_cell.fill = make_fill(GREEN_BG)
    pb_cell.font = Font(name="Calibri", bold=False, color="1A3A1A", size=10)
    pb_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    pb_cell.border = thin_border()
    ws.row_dimensions[body_row].height = 80

    add_sheet_title(ws,
                    "Family Offices — Dutch Bros / Project Alpine Pitch",
                    "Target list for $94.5M entry | 14.3% yield | Permanent Capital | No ROFR exit risk",
                    num_cols)

    freeze_and_filter(ws, freeze_cell="A4",
                      filter_range=f"A3:{get_column_letter(num_cols)}3")

    widths = [28, 16, 20, 22, 34, 24, 10, 60]
    set_col_widths(ws, widths)

    total_rows = 3 + len(data)
    for r in range(4, total_rows + 1):
        ws.row_dimensions[r].height = 48

    return ws, len(data)


def build_index_sheet(wb, sheet_stats):
    """Create a front-page index / legend sheet."""
    ws = wb.create_sheet("INDEX & LEGEND", 0)

    # Main title
    ws.merge_cells("A1:F1")
    t = ws.cell(row=1, column=1,
                value="PE BUYERS MASTER DATABASE — PAULSON INVESTMENT COMPANY")
    t.fill = make_fill(NAVY)
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=15)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    s = ws.cell(row=2, column=1,
                value="Boutique M&A Advisory | QSR · Home Services · Environmental · Specialty · Family Offices · High-Growth")
    s.fill = make_fill("2E5F9E")
    s.font = Font(name="Calibri", italic=True, color=WHITE, size=10)
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Spacer
    ws.row_dimensions[3].height = 8

    # Sheet directory header
    dir_header_cols = ["Sheet", "Description", "# Buyers", "Primary Mandate", "Priority"]
    for i, h in enumerate(dir_header_cols, 1):
        c = ws.cell(row=4, column=i, value=h)
        c.fill = navy_fill()
        c.font = header_font()
        c.alignment = center()
        c.border = thin_border()
    ws.row_dimensions[4].height = 18

    dir_data = []
    for name, count, desc, mandate, priority in sheet_stats:
        dir_data.append([name, desc, count, mandate, priority])

    priority_colors = {
        "CRITICAL": "FF0000",
        "HIGH":     "FF6600",
        "MEDIUM":   "0070C0",
        "SUPPORT":  "595959",
    }

    for r_off, row in enumerate(dir_data):
        r = 5 + r_off
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.border = thin_border()
            cell.alignment = left()
            cell.fill = make_fill("EEF2F7")
            cell.font = body_font(size=10)
            if c_idx == 5:
                color = priority_colors.get(val, "000000")
                cell.font = Font(name="Calibri", bold=True, color=color, size=10)
        ws.row_dimensions[r].height = 20

    # Legend
    ws.row_dimensions[5 + len(dir_data) + 1].height = 10

    leg_start = 5 + len(dir_data) + 2
    ws.merge_cells(start_row=leg_start, start_column=1,
                   end_row=leg_start, end_column=5)
    lh = ws.cell(row=leg_start, column=1, value="EMAIL CONFIDENCE LEGEND")
    lh.fill = navy_fill()
    lh.font = header_font()
    lh.alignment = center()
    ws.row_dimensions[leg_start].height = 18

    legend_items = [
        ("✓ Confirmed",               GREEN_BG, DARK_GREEN,
         "Verified via ZoomInfo, direct reply, or public source. Use directly."),
        ("⚡ High confidence",         AMBER_BG, DARK_AMBER,
         "Format inferred from known pattern. High probability of delivery."),
        ("○ Domain only / call first", GRAY_BG,  DARK_GRAY,
         "Generic inbox only. Call main number first; get named contact before emailing."),
    ]

    for r_off, (label, bg, fg, desc) in enumerate(legend_items):
        r = leg_start + 1 + r_off
        lc = ws.cell(row=r, column=1, value=label)
        lc.fill = make_fill(bg)
        lc.font = Font(name="Calibri", bold=True, color=fg, size=10)
        lc.border = thin_border()
        lc.alignment = center()

        dc = ws.cell(row=r, column=2)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        dc = ws.cell(row=r, column=2, value=desc)
        dc.fill = make_fill(bg)
        dc.font = Font(name="Calibri", bold=False, color=fg, size=10)
        dc.border = thin_border()
        dc.alignment = left()
        ws.row_dimensions[r].height = 22

    # Disclaimer
    disc_row = leg_start + len(legend_items) + 2
    ws.merge_cells(start_row=disc_row, start_column=1,
                   end_row=disc_row, end_column=5)
    disc = ws.cell(row=disc_row, column=1,
                   value=(
                       "DISCLAIMER: This database is confidential and proprietary to Paulson Investment Company. "
                       "Contact data sourced from public filings, ZoomInfo, press releases, and direct verification. "
                       "Fabricated personal emails are NOT included — all personal emails are noted with confidence level. "
                       "Last updated: June 2026."
                   ))
    disc.fill = make_fill("FFF2CC")
    disc.font = Font(name="Calibri", italic=True, color="7D4E00", size=9)
    disc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    disc.border = thin_border()
    ws.row_dimensions[disc_row].height = 40

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12

    return ws


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════

def main():
    output_path = "/home/user/Regular/PE_Master_Database.xlsx"
    wb = openpyxl.Workbook()
    # Remove default sheet
    del wb["Sheet"]

    sheet_stats = []   # (sheet_name, row_count, description, mandate, priority)

    # ── Sheet 1: QSR & Restaurant PE ─────────────────────────────────────────
    ws1, cnt1 = build_standard_sheet(
        wb,
        sheet_name="QSR & Restaurant PE",
        title_text="QSR & Restaurant PE Buyers",
        subtitle_text="Project Alpine (Dutch Bros 35-unit) | $94.5M ask | Outreach: PE firms + Strategic operators",
        headers=QSR_HEADERS,
        data=QSR_DATA,
        conf_col_idx=6,
        col_widths=[26, 20, 20, 18, 34, 24, 16, 10, 14, 52],
    )
    sheet_stats.append((
        "QSR & Restaurant PE", cnt1,
        "PE firms & strategic operators for restaurant/franchise deals",
        "Project Alpine — Dutch Bros 35-unit",
        "CRITICAL"
    ))

    # ── Sheet 2: Home Services PE ─────────────────────────────────────────────
    ws2, cnt2 = build_standard_sheet(
        wb,
        sheet_name="Home Services PE",
        title_text="Home Services PE Buyers",
        subtitle_text="HVAC, plumbing, electrical, home services rollups | LMM to middle market",
        headers=HOME_HEADERS,
        data=HOME_DATA,
        conf_col_idx=5,
        col_widths=[26, 22, 16, 34, 24, 16, 14, 52],
    )
    sheet_stats.append((
        "Home Services PE", cnt2,
        "HVAC, plumbing, electrical, home services rollups",
        "Active home services mandates",
        "HIGH"
    ))

    # ── Sheet 3: Environmental & Waste PE ────────────────────────────────────
    ws3, cnt3 = build_standard_sheet(
        wb,
        sheet_name="Environmental & Waste PE",
        title_text="Environmental & Waste PE Buyers",
        subtitle_text="Fire safety, liquid waste, specialty recycling, environmental services",
        headers=ENV_HEADERS,
        data=ENV_DATA,
        conf_col_idx=5,
        col_widths=[28, 20, 16, 34, 24, 16, 12, 52],
    )
    sheet_stats.append((
        "Environmental & Waste PE", cnt3,
        "Fire safety, liquid waste, specialty recycling, env services",
        "Environmental services mandates",
        "HIGH"
    ))

    # ── Sheet 4: Specialty Services PE ───────────────────────────────────────
    ws4, cnt4 = build_standard_sheet(
        wb,
        sheet_name="Specialty Services PE",
        title_text="Specialty Services PE Buyers",
        subtitle_text="Elevator · Calibration Labs · Funeral Homes · Document Shredding · Insurance Agencies",
        headers=SPEC_HEADERS,
        data=SPEC_DATA,
        conf_col_idx=5,
        col_widths=[28, 16, 20, 34, 24, 16, 12, 52],
    )
    sheet_stats.append((
        "Specialty Services PE", cnt4,
        "Elevator, calibration, funeral homes, shredding, insurance",
        "Silver Tsunami + specialty mandates",
        "HIGH"
    ))

    # ── Sheet 5: Family Offices ───────────────────────────────────────────────
    ws5, cnt5 = build_family_office_sheet(wb, "Family Offices — Alpine")
    sheet_stats.append((
        "Family Offices — Alpine", cnt5,
        "PNW family offices & MFOs for Project Alpine Dutch Bros pitch",
        "Project Alpine — permanent capital buyers",
        "CRITICAL"
    ))

    # ── Sheet 6: High-Growth Sector Buyers ───────────────────────────────────
    ws6, cnt6 = build_standard_sheet(
        wb,
        sheet_name="High-Growth Sectors",
        title_text="High-Growth Sector PE Buyers",
        subtitle_text="Pest Control · Pool Service · MSP/IT · ABA Therapy · Landscaping · Collision Repair",
        headers=HG_HEADERS,
        data=HG_DATA,
        conf_col_idx=5,
        col_widths=[26, 16, 16, 34, 24, 16, 12, 52],
    )
    sheet_stats.append((
        "High-Growth Sectors", cnt6,
        "Pest control, pool, MSP/IT, ABA therapy, landscaping, collision",
        "High-growth services mandates",
        "MEDIUM"
    ))

    # ── Index sheet (prepended at position 0) ────────────────────────────────
    build_index_sheet(wb, sheet_stats)

    wb.save(output_path)
    print(f"\n{'='*60}")
    print(f"  FILE: {output_path}")
    print(f"{'='*60}")
    print(f"  Sheet breakdown:")
    total = 0
    for name, count, *_ in sheet_stats:
        print(f"    {name:<35} {count:>3} buyer rows")
        total += count
    print(f"  {'TOTAL BUYER ROWS':<35} {total:>3}")
    print(f"{'='*60}\n")


if __name__ == "__main__":
    main()
